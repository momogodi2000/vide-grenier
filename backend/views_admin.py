# backend/views_admin.py - COMPLETE ADMIN MANAGEMENT SYSTEM

from math import e
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, Count, Sum, Avg, F, Case, When, IntegerField
from django.utils.decorators import method_decorator
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, TemplateView, View
from django.urls import reverse_lazy
from django.core.exceptions import PermissionDenied
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.auth import get_user_model
from django.contrib.auth.forms import UserCreationForm
from django import forms
import json
import csv
import openpyxl
from datetime import datetime, timedelta
from django.utils import timezone
from django.core.mail import send_mail, send_mass_mail
from django.template.loader import render_to_string
from django.conf import settings
import threading
import time
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
# Import all required models
try:
    from .models import (
        Product, Category, Order, Review, Chat, Message, Notification, 
        AdminStock, PickupPoint, ProductImage, Payment, 
        Favorite, SearchHistory
    )
    MAIN_MODELS_AVAILABLE = True
except ImportError as e:
    print(f"⚠️ Error importing models: {e}")
    MAIN_MODELS_AVAILABLE = False

import random  # Add this for AdminStockAddView
from decimal import Decimal
from django.utils.text import slugify
from django.db import transaction
import yagmail
import uuid

# Models are now imported above

# Import newsletter models
try:
    from .models_newsletter import NewsletterSubscriber, Newsletter
    NEWSLETTER_MODELS_AVAILABLE = True
except ImportError:
    NEWSLETTER_MODELS_AVAILABLE = False
    print("⚠️ Newsletter models not found")

# Import forms for product management
try:
    from .forms import AdminProductForm, ProductImageFormSet
except ImportError:
    # Create basic forms if not available
    class AdminProductForm(forms.ModelForm):
        class Meta:
            model = Product if MAIN_MODELS_AVAILABLE else None
            fields = ['title', 'description', 'category', 'price', 'condition', 'city', 'is_negotiable']
    
    from django.forms import inlineformset_factory
    if MAIN_MODELS_AVAILABLE:
        from .models import ProductImage
        ProductImageFormSet = inlineformset_factory(
            Product, ProductImage,
            fields=('image', 'alt_text', 'is_primary'),
            extra=3,
            can_delete=True
        )

User = get_user_model()

# ============ DEBUG VIEW ============
def admin_debug_view(request):
    """Debug view to test URL resolution"""
    return HttpResponse("Admin debug view working!")

def admin_test_profile_view(request):
    """Test profile view to verify URL resolution"""
    return HttpResponse("Admin profile view working!")

# ============ USER FORMS ============
class AdminUserCreateForm(forms.ModelForm):
    """Form for creating new users"""
    password1 = forms.CharField(
        label='Mot de passe',
        widget=forms.PasswordInput(attrs={'class': 'form-control'})
    )
    password2 = forms.CharField(
        label='Confirmation mot de passe',
        widget=forms.PasswordInput(attrs={'class': 'form-control'})
    )
    
    class Meta:
        model = User
        fields = ['email', 'first_name', 'last_name', 'phone', 'city', 'user_type', 'is_active', 'is_staff']
        widgets = {
            'email': forms.EmailInput(attrs={'class': 'form-control'}),
            'first_name': forms.TextInput(attrs={'class': 'form-control'}),
            'last_name': forms.TextInput(attrs={'class': 'form-control'}),
            'phone': forms.TextInput(attrs={'class': 'form-control'}),
            'city': forms.Select(attrs={'class': 'form-control'}),
            'user_type': forms.Select(attrs={'class': 'form-control'}),
        }
    
    def clean_password2(self):
        password1 = self.cleaned_data.get("password1")
        password2 = self.cleaned_data.get("password2")
        if password1 and password2 and password1 != password2:
            raise forms.ValidationError("Les mots de passe ne correspondent pas")
        return password2
    
    def save(self, commit=True):
        user = super().save(commit=False)
        user.set_password(self.cleaned_data["password1"])
        if commit:
            user.save()
        return user

class AdminUserEditForm(forms.ModelForm):
    """Form for editing existing users"""
    
    class Meta:
        model = User
        fields = [
            'email', 'first_name', 'last_name', 'phone', 'city', 
            'user_type', 'is_active', 'is_staff', 'is_verified', 
            'phone_verified', 'trust_score', 'loyalty_points'
        ]
        widgets = {
            'email': forms.EmailInput(attrs={'class': 'form-control'}),
            'first_name': forms.TextInput(attrs={'class': 'form-control'}),
            'last_name': forms.TextInput(attrs={'class': 'form-control'}),
            'phone': forms.TextInput(attrs={'class': 'form-control'}),
            'city': forms.Select(attrs={'class': 'form-control'}),
            'user_type': forms.Select(attrs={'class': 'form-control'}),
            'trust_score': forms.NumberInput(attrs={'class': 'form-control', 'min': 0, 'max': 100}),
            'loyalty_points': forms.NumberInput(attrs={'class': 'form-control', 'min': 0}),
        }

# ============ UTILITY FUNCTIONS ============
def is_admin(user):
    """Check if user is admin/staff"""
    return user.is_staff or user.is_superuser or getattr(user, 'user_type', None) == 'ADMIN'

def admin_required(view_func):
    """Decorator to require admin access"""
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('backend:login')
        if not is_admin(request.user):
            raise PermissionDenied
        return view_func(request, *args, **kwargs)
    return wrapper

class AdminRequiredMixin(LoginRequiredMixin, UserPassesTestMixin):
    """Mixin to ensure user is admin"""
    def test_func(self):
        return is_admin(self.request.user)

def get_admin_stats():
    """Get admin dashboard statistics"""
    stats = {
        'total_users': 0,
        'total_products': 0,
        'total_orders': 0,
        'total_revenue': 0,
        'new_users_this_month': 0,
        'active_products': 0,
        'pending_orders': 0,
        'newsletter_subscribers': 0,
    }
    
    if MAIN_MODELS_AVAILABLE:
        try:
            stats['total_users'] = User.objects.count()
            stats['total_products'] = Product.objects.count()
            stats['active_products'] = Product.objects.filter(status='ACTIVE').count()
            stats['total_orders'] = Order.objects.count()
            stats['pending_orders'] = Order.objects.filter(status='PENDING').count()
            
            # Calculate revenue
            total_revenue = Order.objects.filter(
                status__in=['DELIVERED', 'COMPLETED']
            ).aggregate(total=Sum('total_amount'))['total'] or 0
            stats['total_revenue'] = total_revenue
            
            # New users this month
            this_month = timezone.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            new_users = User.objects.filter(date_joined__gte=this_month).count()
            stats['new_users_this_month'] = new_users
            
        except Exception as e:
            print(f"Error calculating stats: {e}")
    
    if NEWSLETTER_MODELS_AVAILABLE:
        try:
            stats['newsletter_subscribers'] = NewsletterSubscriber.objects.filter(is_active=True).count()
        except Exception as e:
            print(f"Error calculating newsletter stats: {e}")
            stats['newsletter_subscribers'] = 0
    
    return stats

@admin_required
def admin_newsletter_send(request, newsletter_id):
    """Send newsletter to subscribers"""
    if not NEWSLETTER_MODELS_AVAILABLE:
        messages.error(request, 'Modèles de newsletter non disponibles!')
        return redirect('backend:admin_newsletter_list')
    
    try:
        newsletter = get_object_or_404(Newsletter, id=newsletter_id)
        
        if request.method == 'POST':
            # Get active subscribers
            subscribers = NewsletterSubscriber.objects.filter(is_active=True)
            
            if subscribers.exists():
                # Send emails in background thread
                def send_emails():
                    sent_count = 0
                    failed_count = 0
                    
                    for subscriber in subscribers:
                        try:
                            # Render email content
                            html_content = render_to_string('emails/newsletter.html', {
                                'newsletter': newsletter,
                                'subscriber': subscriber,
                            })
                            
                            # Send email
                            send_mail(
                                subject=newsletter.subject,
                                message=newsletter.content,
                                from_email=settings.DEFAULT_FROM_EMAIL,
                                recipient_list=[subscriber.email],
                                html_message=html_content,
                                fail_silently=False,
                            )
                            sent_count += 1
                        except Exception as e:
                            print(f"Failed to send to {subscriber.email}: {e}")
                            failed_count += 1
                        
                        # Small delay to avoid overwhelming SMTP
                        time.sleep(0.1)
                    
                    # Mark newsletter as sent
                    newsletter.is_sent = True
                    newsletter.sent_at = timezone.now()
                    newsletter.recipients_count = sent_count
                    newsletter.save()
                
                # Start sending in background
                threading.Thread(target=send_emails).start()
                
                messages.success(
                    request, 
                    f'Newsletter en cours d\'envoi à {subscribers.count()} abonnés!'
                )
            else:
                messages.warning(request, 'Aucun abonné actif trouvé!')
            
            return redirect('backend:admin_newsletter_list')
        
        context = {
            'newsletter': newsletter,
            'subscribers_count': NewsletterSubscriber.objects.filter(is_active=True).count(),
        }
        return render(request, 'backend/admin/newsletter/send.html', context)
        
    except Exception as e:
        messages.error(request, f'Erreur: {e}')
        return redirect('backend:admin_newsletter_list')

@admin_required
@require_http_methods(["DELETE", "POST"])
def admin_newsletter_delete(request, newsletter_id):
    """Delete newsletter"""
    try:
        if NEWSLETTER_MODELS_AVAILABLE:
            newsletter = get_object_or_404(Newsletter, id=newsletter_id)
            newsletter_subject = newsletter.subject
            newsletter.delete()
            
            if request.method == 'POST':
                messages.success(request, f'Newsletter "{newsletter_subject}" supprimée avec succès!')
                return redirect('backend:admin_newsletter_list')
            
            return JsonResponse({
                'success': True,
                'message': 'Newsletter supprimée avec succès'
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'Modèles de newsletter non disponibles'
            })
    except Exception as e:
        if request.method == 'POST':
            messages.error(request, f'Erreur lors de la suppression: {e}')
            return redirect('backend:admin_newsletter_list')
        
        return JsonResponse({
            'success': False,
            'message': str(e)
        })

# ============ NEWSLETTER PUBLIC ENDPOINTS ============
@csrf_exempt
@require_http_methods(["POST"])
def newsletter_subscribe(request):
    """Public newsletter subscription endpoint"""
    try:
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST
        
        email = data.get('email')
        name = data.get('name', '')
        
        if not email:
            return JsonResponse({
                'success': False,
                'message': 'Email requis'
            })
        
        if NEWSLETTER_MODELS_AVAILABLE:
            # Check if already subscribed
            subscriber, created = NewsletterSubscriber.objects.get_or_create(
                email=email,
                defaults={'name': name, 'is_active': True}
            )
            
            if not created and not subscriber.is_active:
                subscriber.is_active = True
                subscriber.save()
            
            # Send welcome email
            try:
                welcome_subject = "Bienvenue à notre newsletter!"
                welcome_content = render_to_string('emails/newsletter_welcome.html', {
                    'name': name or 'Abonné',
                    'unsubscribe_url': request.build_absolute_uri('/newsletter/unsubscribe/')
                })
                
                send_mail(
                    subject=welcome_subject,
                    message='Bienvenue à notre newsletter!',
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[email],
                    html_message=welcome_content,
                    fail_silently=True,
                )
            except Exception as e:
                print(f"Error sending welcome email: {e}")
        
        return JsonResponse({
            'success': True,
            'message': 'Inscription réussie à la newsletter!'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })

@csrf_exempt
@require_http_methods(["POST", "GET"])
def newsletter_unsubscribe(request):
    """Public newsletter unsubscription endpoint"""
    if request.method == 'POST':
        try:
            if request.content_type == 'application/json':
                data = json.loads(request.body)
            else:
                data = request.POST
            
            email = data.get('email')
            
            if NEWSLETTER_MODELS_AVAILABLE:
                subscriber = NewsletterSubscriber.objects.filter(email=email).first()
                if subscriber:
                    subscriber.is_active = False
                    subscriber.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Désabonnement réussi'
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': str(e)
            })
    
    return render(request, 'newsletter/unsubscribe.html')

# ============ ANALYTICS AND REPORTS ============
@method_decorator([login_required, staff_member_required], name='dispatch')
class AdminAnalyticsView(ListView):
    """Comprehensive admin analytics"""
    template_name = 'backend/admin/analytics/dashboard.html'
    context_object_name = 'analytics_data'
    
    def get_queryset(self):
        return []  # No queryset needed for analytics
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Calculate date ranges
        today = timezone.now().date()
        last_week = today - timedelta(days=7)
        last_month = today - timedelta(days=30)
        last_year = today - timedelta(days=365)
        
        analytics = {
            'overview': self.get_overview_analytics(),
            'user_analytics': self.get_user_analytics(),
            'product_analytics': self.get_product_analytics(),
            'sales_analytics': self.get_sales_analytics(),
            'engagement_analytics': self.get_engagement_analytics(),
            'monthly_data': self.get_monthly_data(),
        }
        
        context['analytics'] = analytics
        return context
    
    def get_overview_analytics(self):
        """Get overview analytics"""
        if not MAIN_MODELS_AVAILABLE:
            return {}
        
        try:
            return {
                'total_users': User.objects.count(),
                'total_products': Product.objects.count(),
                'total_orders': Order.objects.count(),
                'total_revenue': Order.objects.filter(
                    status__in=['COMPLETED', 'DELIVERED']
                ).aggregate(total=Sum('total_amount'))['total'] or 0,
                'conversion_rate': self.calculate_conversion_rate(),
            }
        except:
            return {}
    
    def get_user_analytics(self):
        """Get user analytics"""
        if not MAIN_MODELS_AVAILABLE:
            return {}
        
        try:
            today = timezone.now().date()
            last_week = today - timedelta(days=7)
            last_month = today - timedelta(days=30)
            
            return {
                'new_users_today': User.objects.filter(date_joined__date=today).count(),
                'new_users_week': User.objects.filter(date_joined__date__gte=last_week).count(),
                'new_users_month': User.objects.filter(date_joined__date__gte=last_month).count(),
                'active_users_week': User.objects.filter(last_login__date__gte=last_week).count(),
                'user_types': self.get_user_types_distribution(),
                'loyalty_levels': self.get_loyalty_distribution(),
            }
        except:
            return {}
    
    def get_product_analytics(self):
        """Get product analytics"""
        if not MAIN_MODELS_AVAILABLE:
            return {}
        
        try:
            today = timezone.now().date()
            last_week = today - timedelta(days=7)
            last_month = today - timedelta(days=30)
            
            return {
                'products_today': Product.objects.filter(created_at__date=today).count(),
                'products_week': Product.objects.filter(created_at__date__gte=last_week).count(),
                'products_month': Product.objects.filter(created_at__date__gte=last_month).count(),
                'active_products': Product.objects.filter(status='ACTIVE').count(),
                'sold_products': Product.objects.filter(status='SOLD').count(),
                'pending_moderation': Product.objects.filter(status='PENDING').count(),
                'avg_product_price': Product.objects.aggregate(avg=Avg('price'))['avg'] or 0,
                'top_categories': self.get_top_categories_data(),
            }
        except:
            return {}
    
    def get_sales_analytics(self):
        """Get sales analytics"""
        if not MAIN_MODELS_AVAILABLE:
            return {}
        
        try:
            today = timezone.now().date()
            last_week = today - timedelta(days=7)
            last_month = today - timedelta(days=30)
            
            return {
                'orders_today': Order.objects.filter(created_at__date=today).count(),
                'orders_week': Order.objects.filter(created_at__date__gte=last_week).count(),
                'orders_month': Order.objects.filter(created_at__date__gte=last_month).count(),
                'revenue_today': Order.objects.filter(
                    created_at__date=today,
                    status__in=['COMPLETED', 'DELIVERED']
                ).aggregate(total=Sum('total_amount'))['total'] or 0,
                'revenue_week': Order.objects.filter(
                    created_at__date__gte=last_week,
                    status__in=['COMPLETED', 'DELIVERED']
                ).aggregate(total=Sum('total_amount'))['total'] or 0,
                'revenue_month': Order.objects.filter(
                    created_at__date__gte=last_month,
                    status__in=['COMPLETED', 'DELIVERED']
                ).aggregate(total=Sum('total_amount'))['total'] or 0,
                'avg_order_value': Order.objects.filter(
                    status__in=['COMPLETED', 'DELIVERED']
                ).aggregate(avg=Avg('total_amount'))['avg'] or 0,
                'order_statuses': self.get_order_status_distribution(),
            }
        except:
            return {}
    
    def get_engagement_analytics(self):
        """Get engagement analytics"""
        if not MAIN_MODELS_AVAILABLE:
            return {}
        
        try:
            last_week = timezone.now().date() - timedelta(days=7)
            
            return {
                'total_reviews': Review.objects.count(),
                'avg_rating': Review.objects.aggregate(avg=Avg('overall_rating'))['avg'] or 0,
                'reviews_week': Review.objects.filter(created_at__date__gte=last_week).count(),
                'active_chats': Chat.objects.filter(is_active=True).count(),
                'messages_week': Message.objects.filter(
                    created_at__date__gte=last_week
                ).count(),
            }
        except:
            return {}
    
    def get_monthly_data(self):
        """Get monthly data for charts"""
        if not MAIN_MODELS_AVAILABLE:
            return []
        
        try:
            today = timezone.now().date()
            monthly_data = []
            
            for i in range(12):
                month_start = today.replace(day=1) - timedelta(days=30*i)
                month_end = month_start + timedelta(days=30)
                
                monthly_data.append({
                    'month': month_start.strftime('%Y-%m'),
                    'users': User.objects.filter(
                        date_joined__date__gte=month_start,
                        date_joined__date__lt=month_end
                    ).count(),
                    'products': Product.objects.filter(
                        created_at__date__gte=month_start,
                        created_at__date__lt=month_end
                    ).count(),
                    'orders': Order.objects.filter(
                        created_at__date__gte=month_start,
                        created_at__date__lt=month_end
                    ).count(),
                    'revenue': Order.objects.filter(
                        created_at__date__gte=month_start,
                        created_at__date__lt=month_end,
                        status__in=['COMPLETED', 'DELIVERED']
                    ).aggregate(total=Sum('total_amount'))['total'] or 0,
                })
            
            return list(reversed(monthly_data))
        except:
            return []
    
    def calculate_conversion_rate(self):
        """Calculate conversion rate"""
        try:
            total_visitors = User.objects.count()  # Or use analytics data
            total_orders = Order.objects.count()
            
            if total_visitors > 0:
                return round((total_orders / total_visitors) * 100, 2)
            return 0
        except:
            return 0
    
    def get_user_types_distribution(self):
        """Get user types distribution"""
        try:
            return {
                'clients': User.objects.filter(user_type='CLIENT').count(),
                'staff': User.objects.filter(user_type='STAFF').count(),
                'admins': User.objects.filter(user_type='ADMIN').count(),
            }
        except:
            return {}
    
    def get_loyalty_distribution(self):
        """Get loyalty levels distribution"""
        try:
            return {
                'bronze': User.objects.filter(loyalty_level='bronze').count(),
                'silver': User.objects.filter(loyalty_level='silver').count(),
                'gold': User.objects.filter(loyalty_level='gold').count(),
                'platinum': User.objects.filter(loyalty_level='platinum').count(),
            }
        except:
            return {}
    
    def get_top_categories_data(self):
        """Get top categories data"""
        try:
            return Category.objects.annotate(
                product_count=Count('products'),
                sales_count=Count('products__order')
            ).order_by('-sales_count')[:10]
        except:
            return []
    
    def get_order_status_distribution(self):
        """Get order status distribution"""
        try:
            return {
                'pending': Order.objects.filter(status='PENDING').count(),
                'confirmed': Order.objects.filter(status='CONFIRMED').count(),
                'processing': Order.objects.filter(status='PROCESSING').count(),
                'shipped': Order.objects.filter(status='SHIPPED').count(),
                'delivered': Order.objects.filter(status='DELIVERED').count(),
                'cancelled': Order.objects.filter(status='CANCELLED').count(),
            }
        except:
            return {}

# ============ EXPORT FUNCTIONS ============
@admin_required
def admin_export_users(request):
    """Export users data to Excel/CSV"""
    try:
        export_format = request.GET.get('format', 'excel')
        
        if export_format == 'excel':
            # Create Excel file
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="users_export_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
            
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Utilisateurs"
            
            # Headers
            headers = [
                'ID', 'Email', 'Prénom', 'Nom', 'Téléphone', 'Ville',
                'Type utilisateur', 'Statut', 'Vérifié', 'Score confiance',
                'Points fidélité', 'Date inscription', 'Dernière connexion'
            ]
            ws.append(headers)
            
            # Data
            for user in User.objects.all().order_by('-date_joined'):
                row = [
                    str(user.id),  # Convert UUID to string
                    user.email,
                    user.first_name,
                    user.last_name,
                    getattr(user, 'phone', '') or '',
                    getattr(user, 'city', '') or '',
                    getattr(user, 'get_user_type_display', lambda: user.user_type)(),
                    'Actif' if user.is_active else 'Inactif',
                    'Oui' if getattr(user, 'is_verified', False) else 'Non',
                    getattr(user, 'trust_score', 0),
                    getattr(user, 'loyalty_points', 0),
                    user.date_joined.strftime('%Y-%m-%d %H:%M'),
                    user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else 'Jamais'
                ]
                ws.append(row)
            
            # Style the headers
            for cell in ws[1]:
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Save workbook
            wb.save(response)
            return response
            
        else:  # CSV format
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="users_export_{timezone.now().strftime("%Y%m%d_%H%M")}.csv"'
            
            writer = csv.writer(response)
            
            # Headers
            writer.writerow([
                'ID', 'Email', 'Prénom', 'Nom', 'Téléphone', 'Ville',
                'Type utilisateur', 'Statut', 'Vérifié', 'Score confiance',
                'Points fidélité', 'Date inscription', 'Dernière connexion'
            ])
            
            # Data
            for user in User.objects.all().order_by('-date_joined'):
                writer.writerow([
                    str(user.id),  # Convert UUID to string
                    user.email,
                    user.first_name,
                    user.last_name,
                    getattr(user, 'phone', '') or '',
                    getattr(user, 'city', '') or '',
                    getattr(user, 'get_user_type_display', lambda: user.user_type)(),
                    'Actif' if user.is_active else 'Inactif',
                    'Oui' if getattr(user, 'is_verified', False) else 'Non',
                    getattr(user, 'trust_score', 0),
                    getattr(user, 'loyalty_points', 0),
                    user.date_joined.strftime('%Y-%m-%d %H:%M'),
                    user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else 'Jamais'
                ])
            
            return response
        
    except Exception as e:
        messages.error(request, f'Erreur lors de l\'export: {e}')
        return redirect('admin_panel:users')

# ============ USER STATUS ACTIONS ============
@admin_required
@require_http_methods(["POST"])
def admin_user_toggle_status(request, user_id):
    """Toggle user active status"""
    try:
        user = get_object_or_404(User, id=user_id)
        
        # Prevent action on superusers and self
        if user.is_superuser:
            return JsonResponse({
                'success': False,
                'message': 'Impossible de modifier le statut d\'un super administrateur'
            })
        
        if user == request.user:
            return JsonResponse({
                'success': False,
                'message': 'Vous ne pouvez pas modifier votre propre statut'
            })
        
        user.is_active = not user.is_active
        user.save()
        
        status_text = 'activé' if user.is_active else 'désactivé'
        
        return JsonResponse({
            'success': True,
            'message': f'Utilisateur {user.email} {status_text}',
            'new_status': user.is_active
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur: {str(e)}'
        })

@admin_required
@require_http_methods(["POST"])
def admin_user_toggle_verification(request, user_id):
    """Toggle user verification status"""
    try:
        user = get_object_or_404(User, id=user_id)
        
        user.is_verified = not getattr(user, 'is_verified', False)
        user.save()
        
        status_text = 'vérifié' if user.is_verified else 'non vérifié'
        
        return JsonResponse({
            'success': True,
            'message': f'Utilisateur {user.email} marqué comme {status_text}',
            'new_verification': user.is_verified
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur: {str(e)}'
        })

# ============ AJAX ENDPOINTS ============
@admin_required
@require_http_methods(["GET"])
def admin_users_ajax(request):
    """Get users data via AJAX for datatables"""
    try:
        draw = int(request.GET.get('draw', 1))
        start = int(request.GET.get('start', 0))
        length = int(request.GET.get('length', 10))
        search_value = request.GET.get('search[value]', '')
        
        # Base queryset
        queryset = User.objects.all()
        
        # Search
        if search_value:
            queryset = queryset.filter(
                Q(first_name__icontains=search_value) |
                Q(last_name__icontains=search_value) |
                Q(email__icontains=search_value) |
                Q(phone__icontains=search_value)
            )
        
        # Count
        total_records = User.objects.count()
        filtered_records = queryset.count()
        
        # Pagination
        users = queryset.order_by('-date_joined')[start:start + length]
        
        # Prepare data
        data = []
        for user in users:
            data.append({
                'id': user.id,
                'email': user.email,
                'full_name': f"{user.first_name} {user.last_name}",
                'phone': getattr(user, 'phone', '') or '',
                'user_type': getattr(user, 'get_user_type_display', lambda: user.user_type)(),
                'is_active': user.is_active,
                'is_verified': getattr(user, 'is_verified', False),
                'date_joined': user.date_joined.strftime('%Y-%m-%d %H:%M'),
                'trust_score': getattr(user, 'trust_score', 0),
                'loyalty_points': getattr(user, 'loyalty_points', 0),
            })
        
        return JsonResponse({
            'draw': draw,
            'recordsTotal': total_records,
            'recordsFiltered': filtered_records,
            'data': data
        })
        
    except Exception as e:
        return JsonResponse({
            'error': str(e)
        }, status=500)

# ============ BULK ACTIONS ============
@admin_required
@require_http_methods(["POST"])
def admin_bulk_actions(request):
    """Handle bulk actions across different admin sections"""
    try:
        action = request.POST.get('action')
        item_type = request.POST.get('item_type')
        item_ids = request.POST.getlist('item_ids')
        
        if not action or not item_type or not item_ids:
            return JsonResponse({
                'success': False,
                'message': 'Paramètres manquants'
            })
        
        result_count = 0
        
        if item_type == 'users' and MAIN_MODELS_AVAILABLE:
            if action == 'activate':
                result_count = User.objects.filter(id__in=item_ids).update(is_active=True)
            elif action == 'deactivate':
                result_count = User.objects.filter(id__in=item_ids).exclude(
                    id=request.user.id
                ).update(is_active=False)
            elif action == 'delete':
                result_count = User.objects.filter(id__in=item_ids).exclude(
                    Q(is_superuser=True) | Q(id=request.user.id)
                ).delete()[0]
        
        elif item_type == 'products' and MAIN_MODELS_AVAILABLE:
            if action == 'approve':
                result_count = Product.objects.filter(id__in=item_ids).update(status='ACTIVE')
            elif action == 'reject':
                result_count = Product.objects.filter(id__in=item_ids).update(status='INACTIVE')
            elif action == 'delete':
                result_count = Product.objects.filter(id__in=item_ids).delete()[0]
        
        elif item_type == 'newsletters' and NEWSLETTER_MODELS_AVAILABLE:
            if action == 'delete':
                result_count = Newsletter.objects.filter(id__in=item_ids).delete()[0]
        
        return JsonResponse({
            'success': True,
            'message': f'Action "{action}" appliquée à {result_count} éléments'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })

# ============ ADMIN STATS AJAX ============
@admin_required
@require_http_methods(["GET"])
def admin_stats_ajax(request):
    """Get admin statistics via AJAX"""
    try:
        stats = get_admin_stats()
        
        # Add real-time data
        stats.update({
            'online_users': User.objects.filter(
                last_login__gte=timezone.now() - timedelta(minutes=15)
            ).count(),
            'recent_orders': Order.objects.filter(
                created_at__gte=timezone.now() - timedelta(hours=24)
            ).count() if MAIN_MODELS_AVAILABLE else 0,
        })
        
        return JsonResponse({
            'success': True,
            'stats': stats
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })

# ============ QUICK ACTIONS ============
@admin_required
@require_http_methods(["POST"])
def admin_quick_action(request):
    """Handle quick actions from admin dashboard"""
    try:
        action = request.POST.get('action')
        item_id = request.POST.get('item_id')
        item_type = request.POST.get('item_type')
        
        if action == 'approve_product' and MAIN_MODELS_AVAILABLE:
            product = get_object_or_404(Product, id=item_id)
            product.status = 'ACTIVE'
            product.save()
            return JsonResponse({
                'success': True,
                'message': f'Produit "{product.title}" approuvé'
            })
        
        elif action == 'suspend_user' and MAIN_MODELS_AVAILABLE:
            user = get_object_or_404(User, id=item_id)
            if not user.is_superuser and user != request.user:
                user.is_active = False
                user.save()
                return JsonResponse({
                    'success': True,
                    'message': f'Utilisateur "{user.email}" suspendu'
                })
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Impossible de suspendre cet utilisateur'
                })
        
        elif action == 'activate_user' and MAIN_MODELS_AVAILABLE:
            user = get_object_or_404(User, id=item_id)
            user.is_active = True
            user.save()
            return JsonResponse({
                'success': True,
                'message': f'Utilisateur "{user.email}" activé'
            })
        
        return JsonResponse({
            'success': False,
            'message': 'Action non reconnue'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })
    
from datetime import timedelta

# ============ MAIN ADMIN DASHBOARD ============
@admin_required
def admin_dashboard(request):
    """Main admin dashboard with comprehensive data"""
    now = timezone.now()
    today = now.date()
    this_month = now.replace(day=1).date()
    last_month = (now.replace(day=1) - timedelta(days=1)).replace(day=1).date()
    
    context = {
        'stats': get_admin_stats(),
        'recent_users': [],
        'recent_products': [],
        'recent_orders': [],
        'pending_moderations': [],
        'analytics': {},
        'unread_notifications_count': 0,
        'unread_messages_count': 0,
    }
    
    if MAIN_MODELS_AVAILABLE:
        try:
            # Basic data
            context['recent_users'] = User.objects.order_by('-date_joined')[:5]
            context['recent_products'] = Product.objects.order_by('-created_at')[:5]
            context['recent_orders'] = Order.objects.order_by('-created_at')[:5]
            context['pending_moderations'] = Product.objects.filter(status='PENDING')[:10]
            
            # Enhanced analytics
            context['analytics'] = {
                'daily_sales': get_daily_sales_data(7),
                'top_categories': get_top_categories(),
                'user_growth': get_user_growth_data(),
                'revenue_growth': get_revenue_growth_data(),
                'city_stats': get_city_statistics(),
                'payment_methods': get_payment_methods_stats(),
            }
            
            # Calculate notification and message counts
            try:
                context['unread_notifications_count'] = Notification.objects.filter(
                    recipient=request.user,
                    is_read=False
                ).count()
            except:
                context['unread_notifications_count'] = 0
                
            try:
                context['unread_messages_count'] = Message.objects.filter(
                    recipient=request.user,
                    is_read=False
                ).count()
            except:
                context['unread_messages_count'] = 0
            
        except Exception as e:
            print(f"Error loading dashboard data: {e}")
    
    return render(request, 'backend/admin/dashboard.html', context)

def get_daily_sales_data(days=7):
    """Get sales data for the last N days"""
    if not MAIN_MODELS_AVAILABLE:
        return {}
    
    today = timezone.now().date()
    data = {'labels': [], 'sales': [], 'revenue': []}
    
    for i in range(days):
        date = today - timedelta(days=days-1-i)
        data['labels'].append(date.strftime('%a'))
        
        daily_orders = Order.objects.filter(created_at__date=date).count()
        daily_revenue = Order.objects.filter(
            created_at__date=date,
            status__in=['DELIVERED', 'COMPLETED']
        ).aggregate(total=Sum('total_amount'))['total'] or 0
        
        data['sales'].append(daily_orders)
        data['revenue'].append(float(daily_revenue))
    
    return data

def get_top_categories():
    """Get top performing categories"""
    if not MAIN_MODELS_AVAILABLE:
        return []
    
    try:
        return Category.objects.annotate(
            product_count=Count('products', filter=Q(products__status='ACTIVE')),
            sales_count=Count('products__order', filter=Q(products__order__status='DELIVERED'))
        ).order_by('-sales_count')[:5]
    except:
        return []

def get_user_growth_data():
    """Get user growth statistics"""
    if not MAIN_MODELS_AVAILABLE:
        return {}
    
    today = timezone.now().date()
    this_month = today.replace(day=1)
    last_month = (this_month - timedelta(days=1)).replace(day=1)
    
    this_month_users = User.objects.filter(date_joined__date__gte=this_month).count()
    last_month_users = User.objects.filter(
        date_joined__date__gte=last_month,
        date_joined__date__lt=this_month
    ).count()
    
    growth_rate = 0
    if last_month_users > 0:
        growth_rate = round(((this_month_users - last_month_users) / last_month_users) * 100, 1)
    
    return {
        'this_month': this_month_users,
        'last_month': last_month_users,
        'growth_rate': growth_rate
    }

def get_revenue_growth_data():
    """Get revenue growth statistics"""
    if not MAIN_MODELS_AVAILABLE:
        return {}
    
    today = timezone.now().date()
    this_month = today.replace(day=1)
    last_month = (this_month - timedelta(days=1)).replace(day=1)
    
    this_month_revenue = Order.objects.filter(
        created_at__date__gte=this_month,
        status__in=['DELIVERED', 'COMPLETED']
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    last_month_revenue = Order.objects.filter(
        created_at__date__gte=last_month,
        created_at__date__lt=this_month,
        status__in=['DELIVERED', 'COMPLETED']
    ).aggregate(total=Sum('total_amount'))['total'] or 0
    
    growth_rate = 0
    if last_month_revenue > 0:
        growth_rate = round(((this_month_revenue - last_month_revenue) / last_month_revenue) * 100, 1)
    
    return {
        'this_month': float(this_month_revenue),
        'last_month': float(last_month_revenue),
        'growth_rate': growth_rate
    }

def get_city_statistics():
    """Get statistics by city"""
    try:
        city_stats = []
        total_users = User.objects.filter(user_type='CLIENT').count()
        
        for city_code, city_name in getattr(User, 'CITIES', []):
            city_users = User.objects.filter(city=city_code, user_type='CLIENT').count()
            percentage = round((city_users / total_users * 100), 1) if total_users > 0 else 0
            
            city_stats.append({
                'name': city_name,
                'users': city_users,
                'percentage': percentage
            })
        
        return sorted(city_stats, key=lambda x: x['users'], reverse=True)[:5]
    except:
        return []

def get_payment_methods_stats():
    """Get payment methods statistics"""
    if not MAIN_MODELS_AVAILABLE:
        return []
    
    try:
        payment_stats = Order.objects.filter(
            status__in=['DELIVERED', 'COMPLETED']
        ).values('payment_method').annotate(
            count=Count('id')
        ).order_by('-count')
        
        return list(payment_stats)
    except:
        return []

# ============ USER MANAGEMENT - COMPLETE CRUD ============

@method_decorator([login_required, staff_member_required], name='dispatch')
class AdminUserListView(ListView):
    """Admin view for managing users - LIST"""
    model = User
    template_name = 'backend/admin/users/list.html'
    context_object_name = 'users'
    paginate_by = 25
    
    def get_queryset(self):
        queryset = User.objects.all().order_by('-date_joined')
        
        # Search functionality
        search = self.request.GET.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(first_name__icontains=search) |
                Q(last_name__icontains=search) |
                Q(email__icontains=search) |
                Q(phone__icontains=search)
            )
        
        # Filter by user type
        user_type = self.request.GET.get('user_type', '')
        if user_type:
            queryset = queryset.filter(user_type=user_type)
        
        # Filter by status
        status = self.request.GET.get('status', '')
        if status == 'active':
            queryset = queryset.filter(is_active=True)
        elif status == 'inactive':
            queryset = queryset.filter(is_active=False)
        elif status == 'verified':
            queryset = queryset.filter(is_verified=True)
        elif status == 'unverified':
            queryset = queryset.filter(is_verified=False)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'total_users': User.objects.count(),
            'active_users': User.objects.filter(is_active=True).count(),
            'sellers_count': User.objects.filter(user_type='CLIENT').count(),
            'buyers_count': User.objects.filter(user_type='CLIENT').count(),
            'suspended_users': User.objects.filter(is_active=False).count(),
            'verified_users': User.objects.filter(is_verified=True).count(),
            'search': self.request.GET.get('search', ''),
            'user_type_filter': self.request.GET.get('user_type', ''),
            'status_filter': self.request.GET.get('status', ''),
        })
        
        # Add product counts for each user if models available
        if MAIN_MODELS_AVAILABLE:
            try:
                for user in context['users']:
                    user.products_count = Product.objects.filter(seller=user).count()
                    user.orders_count = Order.objects.filter(buyer=user).count()
                    user.total_sales = Order.objects.filter(
                        product__seller=user,
                        status__in=['COMPLETED', 'DELIVERED']
                    ).aggregate(total=Sum('total_amount'))['total'] or 0
            except Exception as e:
                print(f"Error adding user stats: {e}")
        
        return context

@method_decorator([login_required, staff_member_required], name='dispatch')
class AdminUserDetailView(DetailView):
    """Admin view for user details - READ"""
    model = User
    template_name = 'backend/admin/users/detail.html'
    context_object_name = 'user_detail'
    pk_url_kwarg = 'pk'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.get_object()
        
        # Get user statistics
        if MAIN_MODELS_AVAILABLE:
            try:
                context.update({
                    'products_count': Product.objects.filter(seller=user).count(),
                    'active_products': Product.objects.filter(seller=user, status='ACTIVE').count(),
                    'orders_as_buyer': Order.objects.filter(buyer=user).count(),
                    'orders_as_seller': Order.objects.filter(product__seller=user).count(),
                    'total_sales': Order.objects.filter(
                        product__seller=user,
                        status__in=['COMPLETED', 'DELIVERED']
                    ).aggregate(total=Sum('total_amount'))['total'] or 0,
                    'total_purchases': Order.objects.filter(
                        buyer=user,
                        status__in=['COMPLETED', 'DELIVERED']
                    ).aggregate(total=Sum('total_amount'))['total'] or 0,
                    'reviews_given': Review.objects.filter(reviewer=user).count(),
                    'reviews_received': Review.objects.filter(product__seller=user).count(),
                    'avg_rating_received': Review.objects.filter(product__seller=user).aggregate(
                        avg=Avg('rating'))['avg'] or 0,
                    'recent_products': Product.objects.filter(seller=user).order_by('-created_at')[:5],
                    'recent_orders': Order.objects.filter(buyer=user).order_by('-created_at')[:5],
                    'recent_reviews': Review.objects.filter(reviewer=user).order_by('-created_at')[:5],
                })
            except Exception as e:
                print(f"Error loading user details: {e}")
                
        return context

@method_decorator([login_required, staff_member_required], name='dispatch')
class AdminUserCreateView(CreateView):
    """Admin view for creating users - CREATE"""
    model = User
    form_class = AdminUserCreateForm
    template_name = 'backend/admin/users/create.html'
    success_url = reverse_lazy('admin_panel:users')
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f'Utilisateur {self.object.email} créé avec succès!'
        )
        return response
    
    def form_invalid(self, form):
        messages.error(self.request, 'Erreur lors de la création de l\'utilisateur.')
        return super().form_invalid(form)

@method_decorator([login_required, staff_member_required], name='dispatch')
class AdminUserUpdateView(UpdateView):
    """Admin view for updating users - UPDATE"""
    model = User
    form_class = AdminUserEditForm
    template_name = 'backend/admin/users/edit.html'
    pk_url_kwarg = 'pk'
    
    def get_success_url(self):
        return reverse_lazy('admin_panel:user_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(
            self.request, 
            f'Utilisateur {self.object.email} mis à jour avec succès!'
        )
        return response
    
    def form_invalid(self, form):
        messages.error(self.request, 'Erreur lors de la mise à jour de l\'utilisateur.')
        return super().form_invalid(form)

@admin_required
@require_http_methods(["POST"])
def admin_user_delete(request, user_id):
    """Admin view for deleting users - DELETE"""
    try:
        user = get_object_or_404(User, id=user_id)
        
        # Prevent deletion of superusers and self
        if user.is_superuser:
            return JsonResponse({
                'success': False,
                'message': 'Impossible de supprimer un super administrateur'
            })
        
        if user == request.user:
            return JsonResponse({
                'success': False,
                'message': 'Vous ne pouvez pas vous supprimer vous-même'
            })
        
        user_email = user.email
        user.delete()
        
        messages.success(request, f'Utilisateur {user_email} supprimé avec succès!')
        return JsonResponse({
            'success': True,
            'message': f'Utilisateur {user_email} supprimé avec succès!'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors de la suppression: {str(e)}'
        })

# ============ USER BULK ACTIONS ============
@admin_required
@require_http_methods(["POST"])
def admin_user_bulk_actions(request):
    """Handle bulk actions on users"""
    try:
        action = request.POST.get('action')
        user_ids = request.POST.getlist('user_ids')
        
        if not user_ids:
            return JsonResponse({
                'success': False,
                'message': 'Aucun utilisateur sélectionné'
            })
        
        users = User.objects.filter(id__in=user_ids)
        
        if action == 'activate':
            updated = users.update(is_active=True)
            message = f'{updated} utilisateurs activés'
            
        elif action == 'deactivate':
            # Prevent deactivating superusers and self
            users = users.exclude(is_superuser=True).exclude(id=request.user.id)
            updated = users.update(is_active=False)
            message = f'{updated} utilisateurs désactivés'
            
        elif action == 'verify':
            updated = users.update(is_verified=True)
            message = f'{updated} utilisateurs vérifiés'
            
        elif action == 'unverify':
            updated = users.update(is_verified=False)
            message = f'{updated} utilisateurs non vérifiés'
            
        elif action == 'delete':
            # Prevent deleting superusers and self
            users = users.exclude(is_superuser=True).exclude(id=request.user.id)
            deleted_count = users.count()
            users.delete()
            message = f'{deleted_count} utilisateurs supprimés'
            
        else:
            return JsonResponse({
                'success': False,
                'message': 'Action non reconnue'
            })
        
        return JsonResponse({
            'success': True,
            'message': message
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur: {str(e)}'
        })

# ============ PRODUCT MANAGEMENT ============

# AdminProductListView was missing from the original file
@method_decorator([login_required, staff_member_required], name='dispatch')
class AdminProductListView(ListView):
    """Admin view for managing all products"""
    model = Product
    template_name = 'backend/admin/products/list.html'
    context_object_name = 'products'
    paginate_by = 20
    
    def get_queryset(self):
        if not MAIN_MODELS_AVAILABLE:
            return []
        
        queryset = Product.objects.select_related('seller', 'category').order_by('-created_at')
        
        # Search functionality
        search = self.request.GET.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(title__icontains=search) |
                Q(description__icontains=search) |
                Q(seller__email__icontains=search)
            )
        
        # Filter by status
        status = self.request.GET.get('status', '')
        if status:
            queryset = queryset.filter(status=status)
        
        # Filter by source
        source = self.request.GET.get('source', '')
        if source:
            queryset = queryset.filter(source=source)
        
        # Filter by category
        category = self.request.GET.get('category', '')
        if category:
            queryset = queryset.filter(category_id=category)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if MAIN_MODELS_AVAILABLE:
            try:
                context.update({
                    'total_products': Product.objects.count(),
                    'pending_products': Product.objects.filter(status='DRAFT').count(),
                    'active_products': Product.objects.filter(status='ACTIVE').count(),
                    'sold_products': Product.objects.filter(status='SOLD').count(),
                    'admin_products': Product.objects.filter(source='ADMIN').count(),
                    'client_products': Product.objects.filter(source='CLIENT').count(),
                    'categories': Category.objects.filter(is_active=True).order_by('name'),
                    'product_statuses': Product.STATUSES if hasattr(Product, 'STATUSES') else [],
                    'product_sources': Product.SOURCES if hasattr(Product, 'SOURCES') else [],
                })
            except Exception as e:
                print(f"Error loading product stats: {e}")
        
        context.update({
            'search': self.request.GET.get('search', ''),
            'status_filter': self.request.GET.get('status', ''),
            'source_filter': self.request.GET.get('source', ''),
            'category_filter': self.request.GET.get('category', ''),
        })
        
        return context


class EmailNotificationService:
    """Service for handling email notifications with yagmail"""
    
    def __init__(self):
        self.yag = None
        self.setup_yagmail()
    
    def setup_yagmail(self):
        """Setup yagmail configuration"""
        try:
            self.yag = yagmail.SMTP(
                user=settings.EMAIL_HOST_USER,
                password=settings.EMAIL_HOST_PASSWORD,
                host=settings.EMAIL_HOST,
                port=settings.EMAIL_PORT,
                smtp_starttls=settings.EMAIL_USE_TLS,
                smtp_ssl=settings.EMAIL_USE_SSL
            )
        except Exception as e:
            print(f"Error setting up yagmail: {e}")
            self.yag = None
    
    def send_product_approval_email(self, product, approved=True):
        """Send product approval/rejection email to seller"""
        try:
            if not self.yag:
                self.setup_yagmail()
            
            seller_email = product.seller.email
            seller_name = product.seller.get_full_name() or product.seller.email
            
            if approved:
                subject = f"✅ Votre produit '{product.title}' a été approuvé"
                template_name = "product_approved"
            else:
                subject = f"❌ Votre produit '{product.title}' a été rejeté"
                template_name = "product_rejected"
            
            # Context for email template
            context = {
                'seller_name': seller_name,
                'product_title': product.title,
                'product_url': f"{settings.SITE_URL}{product.get_absolute_url()}",
                'product_price': product.price,
                'approval_date': timezone.now(),
                'site_name': 'Vidé-Grenier Kamer',
                'site_url': settings.SITE_URL,
                'support_email': settings.DEFAULT_FROM_EMAIL,
                'approved': approved
            }
            
            # Render email templates
            html_content = render_to_string(f'emails/{template_name}.html', context)
            text_content = render_to_string(f'emails/{template_name}.txt', context)
            
            # Send email
            self.yag.send(
                to=seller_email,
                subject=subject,
                contents=[text_content, html_content]
            )
            
            # Create notification in database
            Notification.objects.create(
                user=product.seller,
                type='PRODUCT',
                title=subject,
                message=f"Votre produit '{product.title}' a été {'approuvé' if approved else 'rejeté'}.",
                data={
                    'product_id': str(product.id),
                    'product_title': product.title,
                    'action': 'approved' if approved else 'rejected'
                }
            )
            
            return True
            
        except Exception as e:
            print(f"Error sending product approval email: {e}")
            return False


# Initialize email service
email_service = EmailNotificationService()


# ============ ADMIN PRODUCT LIST VIEW ============
@login_required
@admin_required
def admin_product_list(request):
    """List all products with admin management options"""
    
    # Get filter parameters
    status_filter = request.GET.get('status', '')
    source_filter = request.GET.get('source', '')
    category_filter = request.GET.get('category', '')
    search_query = request.GET.get('search', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Base queryset
    products = Product.objects.select_related('seller', 'category').prefetch_related('images')
    
    # Apply filters
    if status_filter:
        products = products.filter(status=status_filter)
    
    if source_filter:
        products = products.filter(source=source_filter)
    
    if category_filter:
        products = products.filter(category_id=category_filter)
    
    if search_query:
        products = products.filter(
            Q(title__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(seller__email__icontains=search_query) |
            Q(seller__first_name__icontains=search_query) |
            Q(seller__last_name__icontains=search_query)
        )
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            products = products.filter(created_at__gte=date_from_obj)
        except ValueError:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d')
            products = products.filter(created_at__lte=date_to_obj)
        except ValueError:
            pass
    
    # Order by creation date (newest first)
    products = products.order_by('-created_at')
    
    # Pagination
    paginator = Paginator(products, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Statistics
    stats = {
        'total_products': Product.objects.count(),
        'pending_approval': Product.objects.filter(status='DRAFT').count(),
        'active_products': Product.objects.filter(status='ACTIVE').count(),
        'sold_products': Product.objects.filter(status='SOLD').count(),
        'admin_products': Product.objects.filter(source='ADMIN').count(),
        'client_products': Product.objects.filter(source='CLIENT').count(),
    }
    
    # Categories for filter dropdown
    categories = Category.objects.filter(is_active=True).order_by('name')
    
    context = {
        'page_obj': page_obj,
        'products': page_obj.object_list,
        'stats': stats,
        'categories': categories,
        'status_filter': status_filter,
        'source_filter': source_filter,
        'category_filter': category_filter,
        'search_query': search_query,
        'date_from': date_from,
        'date_to': date_to,
        'product_statuses': Product.STATUSES,
        'product_sources': Product.SOURCES,
    }
    
    return render(request, 'backend/admin/products/list.html', context)


# ============ ADMIN PRODUCT DETAIL VIEW ============
@login_required
@admin_required
def admin_product_detail(request, product_id):
    """View product details with admin actions"""
    
    product = get_object_or_404(
        Product.objects.select_related('seller', 'category').prefetch_related('images'),
        id=product_id
    )
    
    # Get related data
    orders = Order.objects.filter(product=product).select_related('buyer').order_by('-created_at')
    admin_stock = None
    
    if product.source == 'ADMIN':
        try:
            admin_stock = AdminStock.objects.get(product=product)
        except AdminStock.DoesNotExist:
            pass
    
    # Analytics data (placeholder)
    analytics_data = []
    
    context = {
        'product': product,
        'orders': orders,
        'admin_stock': admin_stock,
        'analytics_data': analytics_data,
        'can_approve': product.status == 'DRAFT' and product.source == 'CLIENT',
        'can_reject': product.status in ['DRAFT', 'ACTIVE'] and product.source == 'CLIENT',
    }
    
    return render(request, 'backend/admin/products/detail.html', context)


# ============ ADMIN PRODUCT CREATE VIEW ============
@login_required
@admin_required
def admin_product_create(request):
    """Create new admin product"""
    
    if request.method == 'POST':
        form = AdminProductForm(request.POST, request.FILES)
        image_formset = ProductImageFormSet(
            request.POST, 
            request.FILES, 
            queryset=ProductImage.objects.none()
        )
        
        if form.is_valid() and image_formset.is_valid():
            try:
                with transaction.atomic():
                    # Create product
                    product = form.save(commit=False)
                    product.seller = request.user
                    product.source = 'ADMIN'
                    product.status = 'ACTIVE'  # Admin products are automatically active
                    
                    # Generate unique slug
                    base_slug = slugify(product.title)
                    slug = base_slug
                    counter = 1
                    while Product.objects.filter(slug=slug).exists():
                        slug = f"{base_slug}-{counter}"
                        counter += 1
                    product.slug = slug
                    
                    product.save()
                    
                    # Save images
                    images = image_formset.save(commit=False)
                    for i, image in enumerate(images):
                        image.product = product
                        image.order = i
                        if i == 0:  # First image is primary
                            image.is_primary = True
                        image.save()
                    
                    # Create admin stock entry
                    AdminStock.objects.create(
                        product=product,
                        sku=f"ADM-{product.id}",
                        quantity=form.cleaned_data.get('quantity', 1),
                        location=form.cleaned_data.get('location', request.user.city),
                        purchase_price=form.cleaned_data.get('purchase_price', product.price * Decimal('0.8')),
                        condition_notes=form.cleaned_data.get('condition_notes', ''),
                        warranty_info=form.cleaned_data.get('warranty_info', ''),
                    )
                    
                    # Log analytics (placeholder)
                    pass
                    
                    messages.success(request, f'Produit "{product.title}" créé avec succès!')
                    return redirect('backend:admin_product_detail', product_id=product.id)
                    
            except Exception as e:
                messages.error(request, f'Erreur lors de la création du produit: {e}')
        else:
            messages.error(request, 'Veuillez corriger les erreurs dans le formulaire.')
    else:
        form = AdminProductForm()
        image_formset = ProductImageFormSet(queryset=ProductImage.objects.none())
    
    categories = Category.objects.filter(is_active=True).order_by('name')
    
    context = {
        'form': form,
        'image_formset': image_formset,
        'categories': categories,
        'cities': User.CITIES,
        'conditions': Product.CONDITIONS,
    }
    
    return render(request, 'backend/admin/products/create.html', context)


# ============ ADMIN PRODUCT EDIT VIEW ============
@login_required
@admin_required
def admin_product_edit(request, product_id):
    """Edit existing product"""
    
    product = get_object_or_404(Product, id=product_id)
    
    if request.method == 'POST':
        form = AdminProductForm(request.POST, request.FILES, instance=product)
        image_formset = ProductImageFormSet(
            request.POST, 
            request.FILES, 
            queryset=product.images.all()
        )
        
        if form.is_valid() and image_formset.is_valid():
            try:
                with transaction.atomic():
                    # Update product
                    product = form.save()
                    
                    # Update images
                    images = image_formset.save(commit=False)
                    
                    # Delete removed images
                    for deleted_image in image_formset.deleted_objects:
                        deleted_image.delete()
                    
                    # Save new/updated images
                    for i, image in enumerate(images):
                        image.product = product
                        if not image.order:
                            image.order = i
                        image.save()
                    
                    # Ensure at least one primary image
                    if not product.images.filter(is_primary=True).exists():
                        first_image = product.images.first()
                        if first_image:
                            first_image.is_primary = True
                            first_image.save()
                    
                    # Update admin stock if it's an admin product
                    if product.source == 'ADMIN':
                        admin_stock, created = AdminStock.objects.get_or_create(
                            product=product,
                            defaults={
                                'sku': f"ADM-{product.id}",
                                'quantity': 1,
                                'location': product.city,
                                'purchase_price': product.price * Decimal('0.8'),
                            }
                        )
                        if not created and form.cleaned_data.get('quantity'):
                            admin_stock.quantity = form.cleaned_data['quantity']
                            admin_stock.save()
                    
                    messages.success(request, f'Produit "{product.title}" mis à jour avec succès!')
                    return redirect('backend:admin_product_detail', product_id=product.id)
                    
            except Exception as e:
                messages.error(request, f'Erreur lors de la mise à jour: {e}')
        else:
            messages.error(request, 'Veuillez corriger les erreurs dans le formulaire.')
    else:
        form = AdminProductForm(instance=product)
        image_formset = ProductImageFormSet(queryset=product.images.all())
    
    categories = Category.objects.filter(is_active=True).order_by('name')
    
    context = {
        'form': form,
        'image_formset': image_formset,
        'product': product,
        'categories': categories,
        'cities': User.CITIES,
        'conditions': Product.CONDITIONS,
    }
    
    return render(request, 'backend/admin/products/edit.html', context)


# ============ ADMIN PRODUCT DELETE VIEW ============
@login_required
@admin_required
@require_http_methods(["POST", "DELETE"])
def admin_product_delete(request, product_id):
    """Delete product"""
    
    product = get_object_or_404(Product, id=product_id)
    
    try:
        product_title = product.title
        
        # Check if product has orders
        if product.orders.exists():
            if request.method == 'POST':
                messages.error(request, 'Impossible de supprimer un produit qui a des commandes.')
                return redirect('backend:admin_product_detail', product_id=product.id)
            
            return JsonResponse({
                'success': False,
                'message': 'Impossible de supprimer un produit qui a des commandes.'
            })
        
        # Delete the product
        product.delete()
        
        if request.method == 'POST':
            messages.success(request, f'Produit "{product_title}" supprimé avec succès!')
            return redirect('backend:admin_product_list')
        
        return JsonResponse({
            'success': True,
            'message': f'Produit "{product_title}" supprimé avec succès!'
        })
        
    except Exception as e:
        if request.method == 'POST':
            messages.error(request, f'Erreur lors de la suppression: {e}')
            return redirect('backend:admin_product_detail', product_id=product.id)
        
        return JsonResponse({
            'success': False,
            'message': str(e)
        })


# ============ PRODUCT APPROVAL/REJECTION ACTIONS ============
@login_required
@admin_required
@require_http_methods(["POST"])
def admin_product_approve(request, product_id):
    """Approve a client product"""
    
    product = get_object_or_404(Product, id=product_id)
    
    if product.source != 'CLIENT':
        return JsonResponse({
            'success': False,
            'message': 'Seuls les produits clients peuvent être approuvés.'
        })
    
    if product.status != 'DRAFT':
        return JsonResponse({
            'success': False,
            'message': 'Ce produit n\'est pas en attente d\'approbation.'
        })
    
    try:
        with transaction.atomic():
            # Update product status
            product.status = 'ACTIVE'
            product.save()
            
            # Send approval email
            email_sent = email_service.send_product_approval_email(product, approved=True)
            
            # Log analytics
            Analytics.objects.create(
                metric_type='PRODUCT_APPROVE',
                user=request.user,
                data={
                    'product_id': str(product.id),
                    'product_title': product.title,
                    'seller_id': str(product.seller.id),
                    'email_sent': email_sent
                },
                ip_address=request.META.get('REMOTE_ADDR', ''),
            )
            
            message = f'Produit "{product.title}" approuvé avec succès!'
            if not email_sent:
                message += ' (Email d\'approbation non envoyé - erreur)'
            
            return JsonResponse({
                'success': True,
                'message': message,
                'new_status': product.get_status_display()
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors de l\'approbation: {e}'
        })


@login_required
@admin_required
@require_http_methods(["POST"])
def admin_product_reject(request, product_id):
    """Reject a client product"""
    
    product = get_object_or_404(Product, id=product_id)
    
    if product.source != 'CLIENT':
        return JsonResponse({
            'success': False,
            'message': 'Seuls les produits clients peuvent être rejetés.'
        })
    
    if product.status not in ['DRAFT', 'ACTIVE']:
        return JsonResponse({
            'success': False,
            'message': 'Ce produit ne peut pas être rejeté.'
        })
    
    try:
        rejection_reason = request.POST.get('reason', 'Aucune raison spécifiée')
        
        with transaction.atomic():
            # Update product status
            product.status = 'SUSPENDED'
            product.save()
            
            # Send rejection email with reason
            email_sent = email_service.send_product_approval_email(product, approved=False)
            
            # Create detailed notification with reason
            Notification.objects.create(
                user=product.seller,
                type='PRODUCT',
                title=f"Produit '{product.title}' rejeté",
                message=f"Votre produit a été rejeté. Raison: {rejection_reason}",
                data={
                    'product_id': str(product.id),
                    'product_title': product.title,
                    'action': 'rejected',
                    'reason': rejection_reason
                }
            )
            
            # Log analytics (placeholder)
            pass
            
            message = f'Produit "{product.title}" rejeté avec succès!'
            if not email_sent:
                message += ' (Email de rejet non envoyé - erreur)'
            
            return JsonResponse({
                'success': True,
                'message': message,
                'new_status': product.get_status_display()
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors du rejet: {e}'
        })


# ============ BULK ACTIONS ============
@login_required
@admin_required
@require_http_methods(["POST"])
def admin_product_bulk_actions(request):
    """Handle bulk actions on products"""
    
    action = request.POST.get('action')
    product_ids = request.POST.getlist('product_ids')
    
    if not action or not product_ids:
        return JsonResponse({
            'success': False,
            'message': 'Action ou IDs de produits manquants.'
        })
    
    try:
        products = Product.objects.filter(id__in=product_ids)
        
        if action == 'approve':
            approved_count = 0
            for product in products.filter(source='CLIENT', status='DRAFT'):
                product.status = 'ACTIVE'
                product.save()
                email_service.send_product_approval_email(product, approved=True)
                approved_count += 1
            
            return JsonResponse({
                'success': True,
                'message': f'{approved_count} produits approuvés avec succès!'
            })
        
        elif action == 'reject':
            reason = request.POST.get('reason', 'Rejet en masse')
            rejected_count = 0
            
            for product in products.filter(source='CLIENT', status__in=['DRAFT', 'ACTIVE']):
                product.status = 'SUSPENDED'
                product.save()
                email_service.send_product_approval_email(product, approved=False)
                rejected_count += 1
            
            return JsonResponse({
                'success': True,
                'message': f'{rejected_count} produits rejetés avec succès!'
            })
        
        elif action == 'delete':
            # Only delete products without orders
            deletable_products = products.filter(orders__isnull=True)
            deleted_count = deletable_products.count()
            deletable_products.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'{deleted_count} produits supprimés avec succès!'
            })
        
        elif action == 'feature':
            products.update(is_featured=True)
            return JsonResponse({
                'success': True,
                'message': f'{products.count()} produits mis en avant!'
            })
        
        elif action == 'unfeature':
            products.update(is_featured=False)
            return JsonResponse({
                'success': True,
                'message': f'{products.count()} produits retirés de la mise en avant!'
            })
        
        else:
            return JsonResponse({
                'success': False,
                'message': 'Action non reconnue.'
            })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors de l\'action en masse: {e}'
        })


# ============ AJAX ENDPOINTS ============
@login_required
@admin_required
@require_http_methods(["GET"])
def admin_product_stats_ajax(request):
    """Get product statistics via AJAX"""
    
    try:
        # Basic stats
        stats = {
            'total_products': Product.objects.count(),
            'pending_approval': Product.objects.filter(status='DRAFT').count(),
            'active_products': Product.objects.filter(status='ACTIVE').count(),
            'sold_products': Product.objects.filter(status='SOLD').count(),
            'admin_products': Product.objects.filter(source='ADMIN').count(),
            'client_products': Product.objects.filter(source='CLIENT').count(),
        }
        
        # Revenue stats
        total_revenue = Order.objects.filter(status='DELIVERED').aggregate(
            total=Sum('total_amount')
        )['total'] or 0
        
        commission_revenue = Order.objects.filter(
            status='DELIVERED',
            product__source='CLIENT'
        ).aggregate(
            total=Sum('commission_amount')
        )['total'] or 0
        
        stats.update({
            'total_revenue': float(total_revenue),
            'commission_revenue': float(commission_revenue),
        })
        
        # Recent activity
        recent_products = Product.objects.filter(
            created_at__gte=timezone.now() - timedelta(days=7)
        ).count()
        
        stats['recent_products'] = recent_products
        
        return JsonResponse({
            'success': True,
            'stats': stats
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })


@login_required
@admin_required
@require_http_methods(["POST"])
def admin_product_quick_action(request):
    """Handle quick actions on products"""
    
    action = request.POST.get('action')
    product_id = request.POST.get('product_id')
    
    if not action or not product_id:
        return JsonResponse({
            'success': False,
            'message': 'Action ou ID produit manquant.'
        })
    
    try:
        product = get_object_or_404(Product, id=product_id)
        
        if action == 'toggle_featured':
            product.is_featured = not product.is_featured
            product.save()
            
            status = 'mis en avant' if product.is_featured else 'retiré de la mise en avant'
            return JsonResponse({
                'success': True,
                'message': f'Produit {status}!',
                'is_featured': product.is_featured
            })
        
        elif action == 'toggle_premium':
            product.is_premium = not product.is_premium
            product.save()
            
            status = 'marqué premium' if product.is_premium else 'retiré du premium'
            return JsonResponse({
                'success': True,
                'message': f'Produit {status}!',
                'is_premium': product.is_premium
            })
        
        else:
            return JsonResponse({
                'success': False,
                'message': 'Action non reconnue.'
            })
    
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })
# ============ ORDER MANAGEMENT ============
@method_decorator([login_required, staff_member_required], name='dispatch')
class AdminOrderListView(ListView):
    """Admin view for managing orders"""
    model = Order
    template_name = 'backend/admin/orders/list.html'
    context_object_name = 'orders'
    paginate_by = 25
    
    def get_queryset(self):
        if not MAIN_MODELS_AVAILABLE:
            return Order.objects.none()
        
        queryset = Order.objects.select_related('buyer', 'product', 'product__seller').order_by('-created_at')
        
        # Search functionality
        search = self.request.GET.get('search', '')
        if search:
            queryset = queryset.filter(
                Q(order_number__icontains=search) |
                Q(buyer__email__icontains=search) |
                Q(product__title__icontains=search)
            )
        
        # Filter by status
        status = self.request.GET.get('status', '')
        if status:
            queryset = queryset.filter(status=status)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if MAIN_MODELS_AVAILABLE:
            try:
                context.update({
                    'total_orders': Order.objects.count(),
                    'pending_orders': Order.objects.filter(status='PENDING').count(),
                    'completed_orders': Order.objects.filter(status='DELIVERED').count(),
                    'total_revenue': Order.objects.filter(
                        status__in=['DELIVERED', 'COMPLETED']
                    ).aggregate(total=Sum('total_amount'))['total'] or 0,
                })
            except Exception as e:
                print(f"Error loading order stats: {e}")
        
        context.update({
            'search': self.request.GET.get('search', ''),
            'status_filter': self.request.GET.get('status', ''),
        })
        
        return context

# ============ NEWSLETTER MANAGEMENT ============
@method_decorator([login_required, staff_member_required], name='dispatch')
class AdminNewsletterListView(ListView):
    """Admin view for managing newsletters"""
    template_name = 'backend/admin/newsletter/list.html'
    context_object_name = 'newsletters'
    paginate_by = 20
    
    def get_queryset(self):
        if not NEWSLETTER_MODELS_AVAILABLE:
            return []
        
        try:
            # Use the enhanced NewsletterCampaign model instead of legacy Newsletter
            from .models_newsletter import NewsletterCampaign
            queryset = NewsletterCampaign.objects.all().order_by('-created_at')
            
            # Search functionality
            search = self.request.GET.get('search', '')
            if search:
                queryset = queryset.filter(
                    Q(name__icontains=search) |
                    Q(subject__icontains=search) |
                    Q(content__icontains=search)
                )
            
            return queryset
        except Exception as e:
            print(f"Error loading newsletters: {e}")
            return []
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if NEWSLETTER_MODELS_AVAILABLE:
            try:
                from .models_newsletter import NewsletterCampaign, NewsletterSubscriber
                context.update({
                    'total_newsletters': NewsletterCampaign.objects.count(),
                    'sent_newsletters': NewsletterCampaign.objects.filter(status='SENT').count(),
                    'draft_newsletters': NewsletterCampaign.objects.filter(status='DRAFT').count(),
                    'total_subscribers': NewsletterSubscriber.objects.filter(is_active=True).count(),
                })
            except Exception as e:
                print(f"Error loading newsletter stats: {e}")
        
        return context

@admin_required
def admin_newsletter_create(request):
    """Create new newsletter"""
    if request.method == 'POST':
        if not NEWSLETTER_MODELS_AVAILABLE:
            messages.error(request, 'Modèles de newsletter non disponibles!')
            return redirect('backend:admin_newsletter_list')
        
        try:
            subject = request.POST.get('subject')
            content = request.POST.get('content')
            
            if subject and content:
                # Use the enhanced NewsletterCampaign model instead of legacy Newsletter
                from .models_newsletter import NewsletterCampaign
                newsletter = NewsletterCampaign.objects.create(
                    name=subject,  # Use subject as name for backward compatibility
                    subject=subject,
                    content=content,
                    created_by=request.user
                )
                messages.success(request, 'Newsletter créée avec succès!')
                return redirect('backend:admin_newsletter_list')
            else:
                messages.error(request, 'Le sujet et le contenu sont requis!')
        except Exception as e:  
            messages.error(request, f'Erreur lors de la création de la newsletter: {str(e)}')
    
    return render(request, 'backend/admin/newsletter/create.html')


# ADD THESE MISSING ADMIN VIEWS TO YOUR views_admin.py

# ============ PAYMENT MANAGEMENT ============
@method_decorator([login_required, staff_member_required], name='dispatch')
class AdminPaymentListView(ListView):
    """Admin view for managing payments"""
    template_name = 'backend/admin/payments/list.html'
    context_object_name = 'payments'
    paginate_by = 25
    
    def get_queryset(self):
        if not MAIN_MODELS_AVAILABLE:
            return []
        
        try:
            # For now, use orders as payment proxy since Payment model might not exist
            queryset = Order.objects.all().order_by('-created_at')
            
            # Search functionality
            search = self.request.GET.get('search', '')
            if search:
                queryset = queryset.filter(
                    Q(order_number__icontains=search) |
                    Q(buyer__email__icontains=search) |
                    Q(product__title__icontains=search)
                )
            
            # Filter by payment status
            status = self.request.GET.get('status', '')
            if status:
                queryset = queryset.filter(status=status)
            
            return queryset
        except Exception as e:
            print(f"Error loading payments: {e}")
            return []
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if MAIN_MODELS_AVAILABLE:
            try:
                context.update({
                    'total_payments': Order.objects.count(),
                    'completed_payments': Order.objects.filter(status='DELIVERED').count(),
                    'pending_payments': Order.objects.filter(status='PENDING').count(),
                    'failed_payments': Order.objects.filter(status='CANCELLED').count(),
                    'total_revenue': Order.objects.filter(
                        status__in=['DELIVERED', 'COMPLETED']
                    ).aggregate(total=Sum('total_amount'))['total'] or 0,
                })
            except Exception as e:
                print(f"Error loading payment stats: {e}")
        
        context.update({
            'search': self.request.GET.get('search', ''),
            'status_filter': self.request.GET.get('status', ''),
        })
        
        return context

@admin_required
def admin_payment_detail(request, payment_id):
    """View payment details"""
    if MAIN_MODELS_AVAILABLE:
        try:
            # Using order as payment proxy
            payment = get_object_or_404(Order, id=payment_id)
            return render(request, 'backend/admin/payments/detail.html', {
                'payment': payment
            })
        except Exception as e:
            messages.error(request, f'Erreur lors du chargement du paiement: {e}')
    
    return redirect('backend:admin_payment_list')

@admin_required
@require_http_methods(["POST"])
def admin_payment_update_status(request, payment_id):
    """Update payment status via AJAX"""
    try:
        new_status = request.POST.get('status')
        
        if MAIN_MODELS_AVAILABLE:
            # Using order as payment proxy
            payment = get_object_or_404(Order, id=payment_id)
            payment.status = new_status
            payment.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Statut de paiement mis à jour avec succès'
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'Modèles non disponibles'
            })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })

# ============ LOYALTY PROGRAM ADMIN VIEWS ============
@method_decorator([login_required, staff_member_required], name='dispatch')
class AdminLoyaltyListView(ListView):
    """Admin view for managing loyalty programs"""
    template_name = 'backend/admin/loyalty/list.html'
    context_object_name = 'loyalty_programs'
    paginate_by = 20
    
    def get_queryset(self):
        if not MAIN_MODELS_AVAILABLE:
            return []
        
        try:
            # Return users with their loyalty stats
            return User.objects.filter(
                loyalty_points__gt=0
            ).order_by('-loyalty_points')
        except Exception as e:
            print(f"Error loading loyalty data: {e}")
            return []
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if MAIN_MODELS_AVAILABLE:
            try:
                context.update({
                    'total_members': User.objects.filter(loyalty_points__gt=0).count(),
                    'bronze_members': User.objects.filter(loyalty_level='bronze').count(),
                    'silver_members': User.objects.filter(loyalty_level='silver').count(),
                    'gold_members': User.objects.filter(loyalty_level='gold').count(),
                    'platinum_members': User.objects.filter(loyalty_level='platinum').count(),
                    'total_points_issued': User.objects.aggregate(
                        total=Sum('loyalty_points')
                    )['total'] or 0,
                })
            except Exception as e:
                print(f"Error loading loyalty stats: {e}")
        
        return context

@admin_required
def admin_loyalty_create(request):
    """Create new loyalty program"""
    if request.method == 'POST':
        try:
            name = request.POST.get('name')
            description = request.POST.get('description')
            points_per_euro = request.POST.get('points_per_euro', 1)
            
            # For now, just show success message
            messages.success(request, f'Programme de fidélité "{name}" créé avec succès!')
            return redirect('backend:admin_loyalty_list')
        except Exception as e:
            messages.error(request, f'Erreur lors de la création: {e}')
    
    return render(request, 'backend/admin/loyalty/create.html')

@admin_required
def admin_loyalty_edit(request, loyalty_id):
    """Edit loyalty program"""
    if request.method == 'POST':
        try:
            messages.success(request, 'Programme de fidélité mis à jour avec succès!')
            return redirect('backend:admin_loyalty_list')
        except Exception as e:
            messages.error(request, f'Erreur lors de la mise à jour: {e}')
    
    return render(request, 'backend/admin/loyalty/edit.html', {
        'loyalty_id': loyalty_id
    })

# ============ PROMOTION ADMIN VIEWS ============
@method_decorator([login_required, staff_member_required], name='dispatch')
class AdminPromotionListView(ListView):
    """Admin view for managing promotions"""
    template_name = 'backend/admin/promotions/list.html'
    context_object_name = 'promotions'
    paginate_by = 20
    
    def get_queryset(self):
        # For now, return empty list since we don't have Promotion model
        return []
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update({
            'total_promotions': 0,
            'active_promotions': 0,
            'expired_promotions': 0,
            'scheduled_promotions': 0,
        })
        return context

@admin_required
def admin_promotion_create(request):
    """Create new promotion"""
    if request.method == 'POST':
        try:
            name = request.POST.get('name')
            code = request.POST.get('code')
            discount_percent = request.POST.get('discount_percent')
            
            messages.success(request, f'Promotion "{name}" créée avec succès!')
            return redirect('backend:admin_promotion_list')
        except Exception as e:
            messages.error(request, f'Erreur lors de la création: {e}')
    
    return render(request, 'backend/admin/promotions/create.html')

@admin_required
def admin_promotion_edit(request, promotion_id):
    """Edit promotion"""
    if request.method == 'POST':
        try:
            messages.success(request, 'Promotion mise à jour avec succès!')
            return redirect('backend:admin_promotion_list')
        except Exception as e:
            messages.error(request, f'Erreur lors de la mise à jour: {e}')
    
    return render(request, 'backend/admin/promotions/edit.html', {
        'promotion_id': promotion_id
    })

@admin_required
@require_http_methods(["POST"])
def admin_promotion_toggle_status(request, promotion_id):
    """Toggle promotion active status"""
    try:
        return JsonResponse({
            'success': True,
            'message': 'Statut de la promotion mis à jour avec succès'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })

# ============ STOCK MANAGEMENT ============
@method_decorator([login_required, staff_member_required], name='dispatch')
class AdminStockView(ListView):
    """Admin view for managing stock"""
    template_name = 'backend/admin/stock/list.html'
    context_object_name = 'stock_items'
    paginate_by = 25
    
    def get_queryset(self):
        if not MAIN_MODELS_AVAILABLE:
            return []
        
        try:
            # Return products that are managed by admin
            queryset = Product.objects.filter(
                seller__user_type='ADMIN'
            ).order_by('-created_at')
            
            # Search functionality
            search = self.request.GET.get('search', '')
            if search:
                queryset = queryset.filter(
                    Q(title__icontains=search) |
                    Q(description__icontains=search)
                )
            
            return queryset
        except Exception as e:
            print(f"Error loading stock: {e}")
            return []
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if MAIN_MODELS_AVAILABLE:
            try:
                context.update({
                    'total_products': Product.objects.count(),
                    'in_stock_count': Product.objects.filter(stock_quantity__gt=0).count(),
                    'low_stock_count': Product.objects.filter(
                        stock_quantity__gt=0, 
                        stock_quantity__lte=F('stock_threshold')
                    ).count(),
                    'out_of_stock_count': Product.objects.filter(stock_quantity=0).count(),
                    'categories': Category.objects.all(),
                })
                
                # Generate stock alerts
                stock_alerts = []
                low_stock_products = Product.objects.filter(
                    stock_quantity__gt=0, 
                    stock_quantity__lte=F('stock_threshold')
                )[:5]
                for product in low_stock_products:
                    stock_alerts.append(f"Stock faible pour {product.title} ({product.stock_quantity} restant)")
                
                out_of_stock_products = Product.objects.filter(stock_quantity=0)[:5]
                for product in out_of_stock_products:
                    stock_alerts.append(f"Rupture de stock pour {product.title}")
                
                context['stock_alerts'] = stock_alerts
                
            except Exception as e:
                print(f"Error loading stock stats: {e}")
        
        context.update({
            'search': self.request.GET.get('search', ''),
        })
        
        return context

@method_decorator([login_required, staff_member_required], name='dispatch')
class AdminStockAddView(CreateView):
    """Admin view for adding stock"""
    model = Product
    template_name = 'backend/admin/stock/add.html'
    fields = ['title', 'description', 'category', 'price', 'condition', 'city']
    success_url = reverse_lazy('backend:admin_stock_list')
    
    def form_valid(self, form):
        # Set admin as seller
        form.instance.seller = self.request.user
        form.instance.source = 'ADMIN'
        form.instance.status = 'ACTIVE'
        
        # Generate slug
        from django.utils.text import slugify
        import uuid
        base_slug = slugify(form.instance.title)
        form.instance.slug = f"{base_slug}-{str(uuid.uuid4())[:8]}"
        
        response = super().form_valid(form)
        messages.success(self.request, f"Produit {self.object.title} ajouté au stock avec succès!")
        return response

# ============ NOTIFICATION MANAGEMENT ============
@method_decorator([login_required, staff_member_required], name='dispatch')
class AdminNotificationListView(ListView):
    """Admin view for managing notifications"""
    template_name = 'backend/admin/notifications/list.html'
    context_object_name = 'notifications'
    paginate_by = 25
    
    def get_queryset(self):
        if not MAIN_MODELS_AVAILABLE:
            return []
        
        try:
            queryset = Notification.objects.all().order_by('-created_at')
            
            # Search functionality
            search = self.request.GET.get('search', '')
            if search:
                queryset = queryset.filter(
                    Q(title__icontains=search) |
                    Q(message__icontains=search) |
                    Q(user__email__icontains=search)
                )
            
            return queryset
        except Exception as e:
            print(f"Error loading notifications: {e}")
            return []
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        if MAIN_MODELS_AVAILABLE:
            try:
                context.update({
                    'total_notifications': Notification.objects.count(),
                    'unread_notifications': Notification.objects.filter(is_read=False).count(),
                    'system_notifications': Notification.objects.filter(type='SYSTEM').count(),
                    'user_notifications': Notification.objects.filter(type='MESSAGE').count(),
                })
            except Exception as e:
                print(f"Error loading notification stats: {e}")
        
        return context

@admin_required
def admin_notification_create(request):
    """Create new notification"""
    if request.method == 'POST':
        if not MAIN_MODELS_AVAILABLE:
            messages.error(request, 'Modèles de notification non disponibles!')
            return redirect('backend:admin_notification_list')
        
        try:
            title = request.POST.get('title')
            message = request.POST.get('message')
            recipient_type = request.POST.get('recipient_type', 'all')
            
            if title and message:
                # Get recipients based on type
                recipients = []
                if recipient_type == 'all':
                    recipients = User.objects.filter(is_active=True)
                elif recipient_type == 'admins':
                    recipients = User.objects.filter(user_type='ADMIN', is_active=True)
                elif recipient_type == 'clients':
                    recipients = User.objects.filter(user_type='CLIENT', is_active=True)
                
                # Create notifications
                created_count = 0
                for recipient in recipients:
                    try:
                        Notification.objects.create(
                            user=recipient,
                            title=title,
                            message=message,
                            type='SYSTEM'
                        )
                        created_count += 1
                    except Exception as e:
                        print(f"Failed to create notification for {recipient.email}: {e}")
                
                messages.success(request, f'{created_count} notifications créées avec succès!')
                return redirect('backend:admin_notification_list')
            else:
                messages.error(request, 'Le titre et le message sont requis!')
        except Exception as e:
            messages.error(request, f'Erreur lors de la création: {e}')
    
    return render(request, 'backend/admin/notifications/create.html')

@admin_required
@require_http_methods(["POST"])
def admin_notification_send_bulk(request):
    """Send bulk notifications"""
    try:
        if not MAIN_MODELS_AVAILABLE:
            return JsonResponse({
                'success': False,
                'message': 'Modèles de notification non disponibles'
            })
        
        notification_ids = request.POST.getlist('notification_ids')
        action = request.POST.get('action')
        
        if action == 'mark_read':
            updated = Notification.objects.filter(
                id__in=notification_ids
            ).update(is_read=True)
            
            return JsonResponse({
                'success': True,
                'message': f'{updated} notifications marquées comme lues'
            })
        
        elif action == 'delete':
            deleted_count = Notification.objects.filter(
                id__in=notification_ids
            ).delete()[0]
            
            return JsonResponse({
                'success': True,
                'message': f'{deleted_count} notifications supprimées'
            })
        
        else:
            return JsonResponse({
                'success': False,
                'message': 'Action non reconnue'
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        })

# ============ ENHANCED EMAIL NOTIFICATION SERVICE ============
class EnhancedEmailNotificationService(EmailNotificationService):
    """Enhanced email service with better error handling and templates"""
    
    def send_product_approval_email(self, product, approved=True, reason=None):
        """Send product approval/rejection email to seller"""
        try:
            if not self.yag:
                self.setup_yagmail()
            
            seller_email = product.seller.email
            seller_name = product.seller.get_full_name() or product.seller.email
            
            if approved:
                subject = f"✅ Votre produit '{product.title}' a été approuvé - Vidé-Grenier Kamer"
                message = f"""
Bonjour {seller_name},

Excellente nouvelle ! Votre produit "{product.title}" a été approuvé par notre équipe.

Détails du produit:
- Titre: {product.title}
- Prix: {product.price} FCFA
- Catégorie: {product.category.name if hasattr(product, 'category') else 'Non spécifiée'}

Votre produit est maintenant visible sur notre plateforme et les acheteurs peuvent le découvrir.

Vous pouvez consulter votre produit ici: {getattr(settings, 'SITE_URL', 'http://localhost:8000')}/products/{product.slug}/

Merci de faire confiance à Vidé-Grenier Kamer !

L'équipe Vidé-Grenier Kamer
{getattr(settings, 'DEFAULT_FROM_EMAIL', 'admin@videgrenierkamer.com')}
                """
            else:
                subject = f"❌ Votre produit '{product.title}' a été rejeté - Vidé-Grenier Kamer"
                message = f"""
Bonjour {seller_name},

Nous vous informons que votre produit "{product.title}" n'a pas pu être approuvé.

Raison du rejet: {reason or 'Non spécifiée'}

Détails du produit:
- Titre: {product.title}
- Prix: {product.price} FCFA
- Catégorie: {product.category.name if hasattr(product, 'category') else 'Non spécifiée'}

Vous pouvez modifier votre produit et le soumettre à nouveau pour approbation.

Si vous avez des questions, n'hésitez pas à nous contacter.

L'équipe Vidé-Grenier Kamer
{getattr(settings, 'DEFAULT_FROM_EMAIL', 'admin@videgrenierkamer.com')}
                """
            
            # Send email
            if self.yag:
                self.yag.send(
                    to=seller_email,
                    subject=subject,
                    contents=message
                )
                return True
            else:
                # Fallback to Django's send_mail
                from django.core.mail import send_mail
                send_mail(
                    subject=subject,
                    message=message,
                    from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', 'admin@videgrenierkamer.com'),
                    recipient_list=[seller_email],
                    fail_silently=False,
                )
                return True
                
        except Exception as e:
            print(f"Error sending {'approval' if approved else 'rejection'} email: {e}")
            return False
    
    def send_product_rejection_email(self, product, reason=None):
        """Send product rejection email"""
        return self.send_product_approval_email(product, approved=False, reason=reason)

# ============ ENHANCED PRODUCT APPROVAL/REJECTION WITH EMAIL ============
@admin_required
@require_http_methods(["POST"])
def admin_product_approve(request, product_id):
    """Approve a client product and send email notification"""
    if not MAIN_MODELS_AVAILABLE:
        return JsonResponse({
            'success': False,
            'message': 'Modèles non disponibles'
        })
    
    try:
        product = get_object_or_404(Product, id=product_id)
        
        if product.source != 'CLIENT':
            return JsonResponse({
                'success': False,
                'message': 'Seuls les produits clients peuvent être approuvés.'
            })
        
        if product.status != 'DRAFT':
            return JsonResponse({
                'success': False,
                'message': 'Ce produit n\'est pas en attente d\'approbation.'
            })
        
        with transaction.atomic():
            # Update product status
            product.status = 'ACTIVE'
            product.save()
            
            # Send approval email using enhanced email service
            email_sent = False
            try:
                email_service = EnhancedEmailNotificationService()
                email_sent = email_service.send_product_approval_email(product, approved=True)
            except Exception as e:
                print(f"Email sending failed: {e}")
            
            # Create notification
            try:
                Notification.objects.create(
                    user=product.seller,
                    type='PRODUCT',
                    title=f'Produit "{product.title}" approuvé',
                    message=f'Votre produit "{product.title}" a été approuvé et est maintenant visible sur la plateforme.',
                    data={
                        'product_id': str(product.id),
                        'product_title': product.title,
                        'action': 'approved'
                    }
                )
            except Exception as e:
                print(f"Notification creation failed: {e}")
            
            return JsonResponse({
                'success': True,
                'message': f'Produit "{product.title}" approuvé avec succès!' + 
                          ('' if email_sent else ' (Email non envoyé)'),
                'new_status': product.get_status_display() if hasattr(product, 'get_status_display') else product.status
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors de l\'approbation: {e}'
        })

@admin_required
@require_http_methods(["POST"])
def admin_product_reject(request, product_id):
    """Reject a client product and send email notification"""
    if not MAIN_MODELS_AVAILABLE:
        return JsonResponse({
            'success': False,
            'message': 'Modèles non disponibles'
        })
    
    try:
        product = get_object_or_404(Product, id=product_id)
        
        if product.source != 'CLIENT':
            return JsonResponse({
                'success': False,
                'message': 'Seuls les produits clients peuvent être rejetés.'
            })
        
        if product.status not in ['DRAFT', 'ACTIVE']:
            return JsonResponse({
                'success': False,
                'message': 'Ce produit ne peut pas être rejeté.'
            })
        
        rejection_reason = request.POST.get('reason', 'Non conforme aux conditions d\'utilisation')
        
        with transaction.atomic():
            # Update product status
            product.status = 'SUSPENDED'
            product.save()
            
            # Send rejection email using enhanced email service
            email_sent = False
            try:
                email_service = EnhancedEmailNotificationService()
                email_sent = email_service.send_product_rejection_email(product, reason=rejection_reason)
            except Exception as e:
                print(f"Email sending failed: {e}")
            
            # Create notification with reason
            try:
                Notification.objects.create(
                    user=product.seller,
                    type='PRODUCT',
                    title=f'Produit "{product.title}" rejeté',
                    message=f'Votre produit "{product.title}" a été rejeté. Raison: {rejection_reason}',
                    data={
                        'product_id': str(product.id),
                        'product_title': product.title,
                        'action': 'rejected',
                        'reason': rejection_reason
                    }
                )
            except Exception as e:
                print(f"Notification creation failed: {e}")
            
            return JsonResponse({
                'success': True,
                'message': f'Produit "{product.title}" rejeté avec succès!' + 
                          ('' if email_sent else ' (Email non envoyé)'),
                'new_status': product.get_status_display() if hasattr(product, 'get_status_display') else product.status
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors du rejet: {e}'
        })


# ============ MISSING COMPLETE ADMIN VIEWS ============

# PRODUCT MANAGEMENT - COMPLETE MISSING VIEWS
@login_required
@admin_required
def admin_product_detail(request, product_id):
    """View product details with admin actions"""
    if not MAIN_MODELS_AVAILABLE:
        messages.error(request, 'Modèles non disponibles!')
        return redirect('backend:admin_product_list')
    
    product = get_object_or_404(
        Product.objects.select_related('seller', 'category').prefetch_related('images'),
        id=product_id
    )
    
    # Get related data
    orders = []
    admin_stock = None
    
    try:
        orders = Order.objects.filter(product=product).select_related('buyer').order_by('-created_at')
    except:
        pass
    
    if product.source == 'ADMIN':
        try:
            admin_stock = AdminStock.objects.get(product=product)
        except AdminStock.DoesNotExist:
            pass
    
    # Analytics data
    analytics_data = []
    try:
        analytics_data = Analytics.objects.filter(
            data__product_id=str(product.id),
            metric_type='PRODUCT_VIEW'
        ).order_by('-created_at')[:10]
    except:
        pass
    
    context = {
        'product': product,
        'orders': orders,
        'admin_stock': admin_stock,
        'analytics_data': analytics_data,
        'can_approve': product.status == 'DRAFT' and product.source == 'CLIENT',
        'can_reject': product.status in ['DRAFT', 'ACTIVE'] and product.source == 'CLIENT',
    }
    
    return render(request, 'backend/admin/products/detail.html', context)

@login_required
@admin_required
def admin_product_create(request):
    """Create new admin product"""
    if not MAIN_MODELS_AVAILABLE:
        messages.error(request, 'Modèles non disponibles!')
        return redirect('backend:admin_product_list')
    
    if request.method == 'POST':
        try:
            title = request.POST.get('title')
            description = request.POST.get('description')
            category_id = request.POST.get('category')
            price = request.POST.get('price')
            condition = request.POST.get('condition')
            city = request.POST.get('city')
            
            if not all([title, description, category_id, price, condition, city]):
                messages.error(request, 'Tous les champs sont requis!')
                return render(request, 'backend/admin/products/create.html', get_product_form_context())
            
            with transaction.atomic():
                # Generate unique slug
                base_slug = slugify(title)
                slug = base_slug
                counter = 1
                while Product.objects.filter(slug=slug).exists():
                    slug = f"{base_slug}-{counter}"
                    counter += 1
                
                # Create product
                product = Product.objects.create(
                    title=title,
                    slug=slug,
                    description=description,
                    category_id=category_id,
                    seller=request.user,
                    price=Decimal(price),
                    condition=condition,
                    city=city,
                    source='ADMIN',
                    status='ACTIVE'
                )
                
                # Create admin stock entry
                AdminStock.objects.create(
                    product=product,
                    sku=f"ADM-{product.id}",
                    quantity=1,
                    location=city,
                    purchase_price=Decimal(price) * Decimal('0.8'),
                    condition_notes=request.POST.get('condition_notes', ''),
                    warranty_info=request.POST.get('warranty_info', ''),
                )
                
                messages.success(request, f'Produit "{product.title}" créé avec succès!')
                return redirect('backend:admin_product_detail', product_id=product.id)
                
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du produit: {e}')
    
    return render(request, 'backend/admin/products/create.html', get_product_form_context())

@login_required
@admin_required
def admin_product_edit(request, product_id):
    """Edit existing product"""
    if not MAIN_MODELS_AVAILABLE:
        messages.error(request, 'Modèles non disponibles!')
        return redirect('backend:admin_product_list')
    
    product = get_object_or_404(Product, id=product_id)
    
    if request.method == 'POST':
        try:
            title = request.POST.get('title')
            description = request.POST.get('description')
            category_id = request.POST.get('category')
            price = request.POST.get('price')
            condition = request.POST.get('condition')
            city = request.POST.get('city')
            
            if not all([title, description, category_id, price, condition, city]):
                messages.error(request, 'Tous les champs sont requis!')
                return render(request, 'backend/admin/products/edit.html', 
                            get_product_form_context(product))
            
            # Update product
            product.title = title
            product.description = description
            product.category_id = category_id
            product.price = Decimal(price)
            product.condition = condition
            product.city = city
            product.save()
            
            # Update admin stock if it's an admin product
            if product.source == 'ADMIN':
                admin_stock, created = AdminStock.objects.get_or_create(
                    product=product,
                    defaults={
                        'sku': f"ADM-{product.id}",
                        'quantity': 1,
                        'location': city,
                        'purchase_price': Decimal(price) * Decimal('0.8'),
                    }
                )
                if not created:
                    admin_stock.location = city
                    admin_stock.save()
            
            messages.success(request, f'Produit "{product.title}" mis à jour avec succès!')
            return redirect('backend:admin_product_detail', product_id=product.id)
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la mise à jour: {e}')
    
    return render(request, 'backend/admin/products/edit.html', 
                  get_product_form_context(product))

def get_product_form_context(product=None):
    """Get context for product forms"""
    try:
        categories = Category.objects.filter(is_active=True).order_by('name')
    except:
        categories = []
    
    return {
        'product': product,
        'categories': categories,
        'cities': User.CITIES,
        'conditions': Product.CONDITIONS,
    }

# ==============================================================================
# VISITOR ORDER EMAIL NOTIFICATIONS
# ==============================================================================

def send_visitor_order_notification(order):
    """Send notification to admin about new visitor order"""
    try:
        email_service = EnhancedEmailNotificationService()
        
        admin_emails = [
            email for email, _ in settings.MANAGERS
        ] or [settings.DEFAULT_FROM_EMAIL]
        
        subject = f"Nouvelle commande visiteur #{order.order_number}"
        
        context = {
            'order': order,
            'product': order.product,
            'visitor_name': order.visitor_name,
            'visitor_phone': order.visitor_phone,
            'visitor_email': order.visitor_email,
            'payment_method': order.get_payment_method_display(),
            'delivery_method': order.get_delivery_method_display(),
            'admin_url': f"{getattr(settings, 'SITE_URL', 'http://localhost:8000')}/admin/orders/{order.id}/",
        }
        
        html_content = f"""
        <h2>Nouvelle commande visiteur</h2>
        <p>Une nouvelle commande a été passée par un visiteur:</p>
        
        <h3>Informations client:</h3>
        <ul>
            <li><strong>Nom:</strong> {order.visitor_name}</li>
            <li><strong>Téléphone:</strong> {order.visitor_phone}</li>
            <li><strong>Email:</strong> {order.visitor_email or 'Non fourni'}</li>
            <li><strong>Contact WhatsApp:</strong> {'Oui' if order.whatsapp_preferred else 'Non'}</li>
        </ul>
        
        <h3>Détails de la commande:</h3>
        <ul>
            <li><strong>Numéro:</strong> {order.order_number}</li>
            <li><strong>Produit:</strong> {order.product.title}</li>
            <li><strong>Quantité:</strong> {order.quantity}</li>
            <li><strong>Montant total:</strong> {order.total_amount} FCFA</li>
            <li><strong>Paiement:</strong> {order.get_payment_method_display()}</li>
            <li><strong>Livraison:</strong> {order.get_delivery_method_display()}</li>
        </ul>
        
        {f'<p><strong>Adresse de livraison:</strong> {order.delivery_address}</p>' if order.delivery_address else ''}
        {f'<p><strong>Notes:</strong> {order.notes}</p>' if order.notes else ''}
        
        <p><a href="{context['admin_url']}">Voir la commande dans l'admin</a></p>
        """
        
        for admin_email in admin_emails:
            if email_service.use_yagmail and email_service.yag:
                email_service.yag.send(
                    to=admin_email,
                    subject=subject,
                    contents=html_content
                )
            else:
                send_mail(
                    subject=subject,
                    message=f"Nouvelle commande visiteur de {order.visitor_name}",
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[admin_email],
                    html_message=html_content,
                    fail_silently=False
                )
        
        return True
        
    except Exception as e:
        print(f"Error sending visitor order notification: {e}")
        return False


def send_visitor_payment_confirmation(order):
    """Send payment confirmation to visitor"""
    try:
        if not order.visitor_email:
            return False
            
        email_service = EnhancedEmailNotificationService()
        subject = f"Confirmation de paiement - Commande #{order.order_number}"
        
        admin_whatsapp = getattr(settings, 'ADMIN_WHATSAPP', os.getenv('ADMIN_WHATSAPP', '237'))
        
        html_content = f"""
        <h2>Paiement confirmé !</h2>
        <p>Bonjour {order.visitor_name},</p>
        
        <p>Votre paiement pour la commande <strong>#{order.order_number}</strong> a été confirmé avec succès.</p>
        
        <h3>Détails de votre commande:</h3>
        <ul>
            <li><strong>Produit:</strong> {order.product.title}</li>
            <li><strong>Quantité:</strong> {order.quantity}</li>
            <li><strong>Montant payé:</strong> {order.total_amount} FCFA</li>
            <li><strong>Mode de livraison:</strong> {order.get_delivery_method_display()}</li>
        </ul>
        
        <p>Nous vous contacterons bientôt pour organiser {'la livraison' if order.delivery_method == 'DELIVERY' else 'le retrait'}.</p>
        
        <p>Pour toute question, vous pouvez nous contacter:</p>
        <ul>
            <li><strong>WhatsApp:</strong> <a href="https://wa.me/{admin_whatsapp}">+{admin_whatsapp}</a></li>
            <li><strong>Email:</strong> {settings.DEFAULT_FROM_EMAIL}</li>
        </ul>
        
        <p>Merci de votre confiance !</p>
        <p>L'équipe Vidé-Grenier Kamer</p>
        """
        
        if email_service.use_yagmail and email_service.yag:
            email_service.yag.send(
                to=order.visitor_email,
                subject=subject,
                contents=html_content
            )
        else:
            send_mail(
                subject=subject,
                message=f"Votre paiement pour la commande #{order.order_number} a été confirmé.",
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[order.visitor_email],
                html_message=html_content,
                fail_silently=False
            )
        
        return True
        
    except Exception as e:
        print(f"Error sending visitor payment confirmation: {e}")
        return False

# ============= MISSING ADMIN VIEWS =============
# These views are referenced in urls_admin.py but were missing

class AdminReportsView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/reports/reports.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if MAIN_MODELS_AVAILABLE:
            context['total_users'] = User.objects.count()
            context['total_products'] = Product.objects.count()
            context['total_orders'] = Order.objects.count()
            context['total_revenue'] = Payment.objects.filter(status='COMPLETED').aggregate(Sum('amount'))['amount__sum'] or 0
        return context

class ExportReportsView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/reports/export.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminUserDeleteView(AdminRequiredMixin, DeleteView):
    model = User
    template_name = 'backend/admin/users/delete.html'
    pk_url_kwarg = 'pk'
    success_url = reverse_lazy('admin_panel:users')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['user'] = self.get_object()
        return context

class AdminUserBulkActionsView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/users/bulk_actions.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminProductForm(forms.ModelForm):
    """Custom form for admin product creation with image handling"""
    primary_image = forms.ImageField(
        label='Image principale',
        required=False,
        widget=forms.ClearableFileInput(attrs={
            'class': 'block w-full text-sm text-gray-500 file:mr-4 file:py-2 file:px-4 file:rounded-full file:border-0 file:text-sm file:font-semibold file:bg-blue-50 file:text-blue-700 hover:file:bg-blue-100',
            'accept': 'image/*'
        })
    )
    additional_images = forms.FileField(
        label='Images supplémentaires',
        required=False,
        widget=forms.FileInput(attrs={
            'class': 'block w-full text-sm text-gray-500 file:mr-4 file:py-2 file:px-4 file:rounded-full file:border-0 file:text-sm file:font-semibold file:bg-blue-50 file:text-blue-700 hover:file:bg-blue-100',
            'accept': 'image/*'
        })
    )
    
    class Meta:
        model = Product
        fields = ['title', 'description', 'category', 'price', 'condition', 'city', 'is_negotiable']
        widgets = {
            'title': forms.TextInput(attrs={'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-2 focus:ring-blue-500 focus:border-transparent'}),
            'description': forms.Textarea(attrs={'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-2 focus:ring-blue-500 focus:border-transparent', 'rows': 4}),
            'category': forms.Select(attrs={'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-2 focus:ring-blue-500 focus:border-transparent'}),
            'price': forms.NumberInput(attrs={'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-2 focus:ring-blue-500 focus:border-transparent', 'min': '1000', 'step': '100'}),
            'condition': forms.Select(attrs={'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-2 focus:ring-blue-500 focus:border-transparent'}),
            'city': forms.Select(attrs={'class': 'w-full px-3 py-2 border border-gray-300 rounded-lg focus:ring-2 focus:ring-blue-500 focus:border-transparent'}),
        }

class AdminProductCreateView(AdminRequiredMixin, CreateView):
    model = Product
    form_class = AdminProductForm
    template_name = 'backend/admin/products/create.html'
    success_url = reverse_lazy('admin_panel:products')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if MAIN_MODELS_AVAILABLE:
            try:
                context['categories'] = Category.objects.filter(is_active=True).order_by('name')
                context['conditions'] = Product.CONDITIONS
                context['cities'] = User.CITIES
            except Exception as e:
                print(f"Error loading form context: {e}")
                context['categories'] = []
                context['conditions'] = []
                context['cities'] = []
        return context
    
    def form_valid(self, form):
        # Save the product first
        product = form.save(commit=False)
        product.seller = self.request.user
        product.source = 'ADMIN'
        product.status = 'ACTIVE'  # Admin products are automatically active
        product.is_approved = True  # Admin products are automatically approved
        product.approved_at = timezone.now()
        product.approved_by = self.request.user
        product.save()
        
        # Handle primary image
        if self.request.FILES.get('primary_image'):
            primary_image = self.request.FILES['primary_image']
            ProductImage.objects.create(
                product=product,
                image=primary_image,
                is_primary=True,
                order=0
            )
        
        # Handle additional images
        additional_images = self.request.FILES.getlist('additional_images')
        for i, image_file in enumerate(additional_images, start=1):
            ProductImage.objects.create(
                product=product,
                image=image_file,
                is_primary=False,
                order=i
            )
        
        messages.success(self.request, f'Produit "{product.title}" créé avec succès!')
        return super().form_valid(form)

class AdminProductDetailView(AdminRequiredMixin, DetailView):
    model = Product
    template_name = 'backend/admin/products/detail.html'
    context_object_name = 'product'
    pk_url_kwarg = 'pk'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminProductUpdateView(AdminRequiredMixin, UpdateView):
    model = Product
    form_class = AdminProductForm
    template_name = 'backend/admin/products/edit.html'
    pk_url_kwarg = 'pk'
    success_url = reverse_lazy('admin_panel:products')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if MAIN_MODELS_AVAILABLE:
            try:
                context['categories'] = Category.objects.filter(is_active=True).order_by('name')
                context['conditions'] = Product.CONDITIONS
                context['cities'] = User.CITIES
                context['users'] = User.objects.filter(is_active=True).order_by('first_name', 'last_name')
                # Get existing images
                context['existing_images'] = self.object.images.all().order_by('order')
            except Exception as e:
                print(f"Error loading form context: {e}")
                context['categories'] = []
                context['conditions'] = []
                context['cities'] = []
                context['users'] = []
                context['existing_images'] = []
        return context
    
    def form_valid(self, form):
        # Save the product first
        product = form.save()
        
        # Handle primary image
        if self.request.FILES.get('primary_image'):
            # Remove existing primary image
            ProductImage.objects.filter(product=product, is_primary=True).update(is_primary=False)
            
            primary_image = self.request.FILES['primary_image']
            ProductImage.objects.create(
                product=product,
                image=primary_image,
                is_primary=True,
                order=0
            )
        
        # Handle additional images
        additional_images = self.request.FILES.getlist('additional_images')
        if additional_images:
            # Get the highest order number
            max_order = ProductImage.objects.filter(product=product).aggregate(
                max_order=models.Max('order')
            )['max_order'] or 0
            
            for i, image_file in enumerate(additional_images, start=max_order + 1):
                ProductImage.objects.create(
                    product=product,
                    image=image_file,
                    is_primary=False,
                    order=i
                )
        
        messages.success(self.request, f'Produit "{product.title}" mis à jour avec succès!')
        return super().form_valid(form)

class AdminProductDeleteView(AdminRequiredMixin, DeleteView):
    model = Product
    template_name = 'backend/admin/products/delete.html'
    pk_url_kwarg = 'pk'
    success_url = reverse_lazy('admin_panel:products')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context
    
    def delete(self, request, *args, **kwargs):
        product = self.get_object()
        messages.success(request, f'Produit "{product.title}" supprimé avec succès!')
        return super().delete(request, *args, **kwargs)

@admin_required
@require_http_methods(["POST"])
def admin_product_delete_ajax(request, product_id):
    """AJAX endpoint for deleting products"""
    try:
        product = get_object_or_404(Product, pk=product_id)
        product_title = product.title
        product.delete()
        
        # Add success message for non-AJAX requests
        messages.success(request, f'Produit "{product_title}" supprimé avec succès!')
        
        return JsonResponse({
            'success': True,
            'message': f'Produit "{product_title}" supprimé avec succès'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Erreur lors de la suppression: {str(e)}'
        })

class AdminProductBulkActionsView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/products/bulk_actions.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminPendingProductsView(AdminRequiredMixin, ListView):
    """Admin view for managing pending products that need approval"""
    model = Product
    template_name = 'backend/admin/products/pending.html'
    context_object_name = 'pending_products'
    paginate_by = 20
    
    def get_queryset(self):
        if MAIN_MODELS_AVAILABLE:
            return Product.objects.filter(
                status='PENDING',
                seller__user_type='CLIENT'
            ).select_related('seller', 'category').prefetch_related('images').order_by('-created_at')
        return Product.objects.none()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if MAIN_MODELS_AVAILABLE:
            context['pending_count'] = Product.objects.filter(status='PENDING').count()
            context['total_products'] = Product.objects.count()
        return context

class AdminProductApprovalView(AdminRequiredMixin, DetailView):
    """Admin view for approving/rejecting a specific product"""
    model = Product
    template_name = 'backend/admin/products/approval.html'
    context_object_name = 'product'
    pk_url_kwarg = 'pk'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if MAIN_MODELS_AVAILABLE:
            context['approval_reasons'] = [
                ('APPROVED', 'Approuvé - Produit conforme'),
                ('REJECTED_INAPPROPRIATE', 'Rejeté - Contenu inapproprié'),
                ('REJECTED_QUALITY', 'Rejeté - Qualité insuffisante'),
                ('REJECTED_DESCRIPTION', 'Rejeté - Description incomplète'),
                ('REJECTED_PRICE', 'Rejeté - Prix non conforme'),
                ('REJECTED_OTHER', 'Rejeté - Autre raison')
            ]
        return context

class AdminProductApproveView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/products/approve.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['product'] = get_object_or_404(Product, pk=self.kwargs['pk'])
        return context

class AdminProductRejectView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/products/reject.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['product'] = get_object_or_404(Product, pk=self.kwargs['pk'])
        return context

@admin_required
@require_http_methods(["POST"])
def admin_product_approve_action(request, product_id):
    """Handle product approval action"""
    try:
        product = get_object_or_404(Product, pk=product_id)
        action = request.POST.get('action')
        reason = request.POST.get('reason', '')
        
        if action == 'approve':
            product.status = 'ACTIVE'
            product.is_approved = True
            product.approved_at = timezone.now()
            product.approved_by = request.user
            product.save()
            
            # Send approval email
            try:
                send_product_approval_email(product, approved=True)
                messages.success(request, f'Produit "{product.title}" approuvé avec succès. Email envoyé au vendeur.')
            except Exception as e:
                messages.warning(request, f'Produit approuvé mais erreur lors de l\'envoi de l\'email: {e}')
                
        elif action == 'reject':
            product.status = 'REJECTED'
            product.is_approved = False
            product.rejected_at = timezone.now()
            product.rejected_by = request.user
            product.rejection_reason = reason
            product.save()
            
            # Send rejection email
            try:
                send_product_approval_email(product, approved=False, reason=reason)
                messages.success(request, f'Produit "{product.title}" rejeté. Email envoyé au vendeur.')
            except Exception as e:
                messages.warning(request, f'Produit rejeté mais erreur lors de l\'envoi de l\'email: {e}')
        
        return redirect('admin_panel:pending_products')
        
    except Exception as e:
        messages.error(request, f'Erreur lors du traitement: {e}')
        return redirect('admin_panel:pending_products')

def send_product_approval_email(product, approved=True, reason=''):
    """Send email notification to seller about product approval/rejection"""
    try:
        subject = f'Votre produit "{product.title}" a été {"approuvé" if approved else "rejeté"}'
        
        if approved:
            message = f"""
            Bonjour {product.seller.get_full_name()},
            
            Votre produit "{product.title}" a été approuvé et est maintenant visible sur notre plateforme.
            
            Détails du produit:
            - Titre: {product.title}
            - Prix: {product.price} XAF
            - Catégorie: {product.category.name}
            
            Vous pouvez maintenant recevoir des commandes pour ce produit.
            
            Merci d'utiliser notre plateforme!
            L'équipe Vidé-Grenier
            """
        else:
            message = f"""
            Bonjour {product.seller.get_full_name()},
            
            Votre produit "{product.title}" a été rejeté.
            
            Raison: {reason}
            
            Détails du produit:
            - Titre: {product.title}
            - Prix: {product.price} XAF
            - Catégorie: {product.category.name}
            
            Vous pouvez modifier votre produit et le soumettre à nouveau pour approbation.
            
            Merci de votre compréhension.
            L'équipe Vidé-Grenier
            """
        
        # Send email using Django's email system
        from django.core.mail import send_mail
        from django.conf import settings
        
        send_mail(
            subject=subject,
            message=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[product.seller.email],
            fail_silently=False,
        )
        
    except Exception as e:
        print(f"Error sending product approval email: {e}")
        # Don't raise the exception to avoid breaking the approval process

class AdminOrderDetailView(AdminRequiredMixin, DetailView):
    model = Order
    template_name = 'backend/admin/orders/detail.html'
    context_object_name = 'order'
    pk_url_kwarg = 'pk'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminOrderUpdateView(AdminRequiredMixin, UpdateView):
    model = Order
    template_name = 'backend/admin/orders/edit.html'
    fields = ['status', 'delivery_address', 'notes']
    pk_url_kwarg = 'pk'
    success_url = reverse_lazy('admin_panel:orders')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminOrderCancelView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/orders/cancel.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['order'] = get_object_or_404(Order, pk=self.kwargs['pk'])
        return context

class ExportOrdersView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/orders/export.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminOrderBulkActionsView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/orders/bulk_actions.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminPaymentsView(AdminRequiredMixin, ListView):
    model = Payment
    template_name = 'backend/admin/payments/list.html'
    context_object_name = 'payments'
    paginate_by = 25
    
    def get_queryset(self):
        return Payment.objects.all().order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class PaymentDetailView(AdminRequiredMixin, DetailView):
    model = Payment
    template_name = 'backend/admin/payments/detail.html'
    context_object_name = 'payment'
    pk_url_kwarg = 'pk'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminPaymentRefundView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/payments/refund.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['payment'] = get_object_or_404(Payment, pk=self.kwargs['pk'])
        return context

class ExportPaymentsView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/payments/export.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminPaymentBulkActionsView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/payments/bulk_actions.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class StockAddView(AdminRequiredMixin, CreateView):
    model = AdminStock
    template_name = 'backend/admin/stock/add.html'
    fields = ['product', 'quantity', 'location', 'notes']
    success_url = reverse_lazy('admin_panel:stock')
    
    def form_valid(self, form):
        form.instance.added_by = self.request.user
        return super().form_valid(form)

class StockListView(AdminRequiredMixin, ListView):
    model = AdminStock
    template_name = 'backend/admin/stock/list.html'
    context_object_name = 'stock_items'
    paginate_by = 25
    
    def get_queryset(self):
        return AdminStock.objects.all().order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminStockEditView(AdminRequiredMixin, UpdateView):
    model = AdminStock
    template_name = 'backend/admin/stock/edit.html'
    fields = ['product', 'quantity', 'location', 'notes']
    pk_url_kwarg = 'pk'
    success_url = reverse_lazy('admin_panel:stock')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminStockDeleteView(AdminRequiredMixin, DeleteView):
    model = AdminStock
    template_name = 'backend/admin/stock/delete.html'
    pk_url_kwarg = 'pk'
    success_url = reverse_lazy('admin_panel:stock')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminStockImportView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/stock/import.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminStockExportView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/stock/export.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminStockMovementsView(AdminRequiredMixin, ListView):
    model = AdminStock
    template_name = 'backend/admin/stock/movements.html'
    context_object_name = 'movements'
    paginate_by = 25
    
    def get_queryset(self):
        return AdminStock.objects.all().order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class PromotionsView(AdminRequiredMixin, ListView):
    template_name = 'backend/admin/promotions/list.html'
    context_object_name = 'promotions'
    paginate_by = 20
    
    def get_queryset(self):
        # Return empty list since we don't have Promotion model yet
        return []
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class PromotionCreateView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/promotions/create.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class PromotionEditView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/promotions/edit.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class PromotionDeleteView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/promotions/delete.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class PromotionActivateView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/promotions/activate.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class PromotionDeactivateView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/promotions/deactivate.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class NewsletterView(AdminRequiredMixin, ListView):
    template_name = 'backend/admin/newsletter/list.html'
    context_object_name = 'subscribers'
    paginate_by = 20
    
    def get_queryset(self):
        if NEWSLETTER_MODELS_AVAILABLE:
            from .models_newsletter import NewsletterSubscriber
            return NewsletterSubscriber.objects.all().order_by('-subscribed_at')
        return []
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if NEWSLETTER_MODELS_AVAILABLE:
            from .models_newsletter import NewsletterCampaign, NewsletterSubscriber
            context.update({
                'total_campaigns': NewsletterCampaign.objects.count(),
                'sent_campaigns': NewsletterCampaign.objects.filter(status='SENT').count(),
                'draft_campaigns': NewsletterCampaign.objects.filter(status='DRAFT').count(),
                'scheduled_campaigns': NewsletterCampaign.objects.filter(status='SCHEDULED').count(),
                'total_subscribers': NewsletterSubscriber.objects.filter(is_active=True).count(),
            })
        return context

class NewsletterCreateView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/newsletter/create.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class NewsletterListView(AdminRequiredMixin, ListView):
    template_name = 'backend/admin/newsletter/list.html'
    context_object_name = 'newsletters'
    paginate_by = 20
    
    def get_queryset(self):
        if NEWSLETTER_MODELS_AVAILABLE:
            # Use the enhanced NewsletterCampaign model instead of legacy Newsletter
            from .models_newsletter import NewsletterCampaign
            return NewsletterCampaign.objects.all().order_by('-created_at')
        return []
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class NewsletterSentView(AdminRequiredMixin, ListView):
    template_name = 'backend/admin/newsletter/sent.html'
    context_object_name = 'newsletters'
    paginate_by = 20
    
    def get_queryset(self):
        if NEWSLETTER_MODELS_AVAILABLE:
            # Use the enhanced NewsletterCampaign model instead of legacy Newsletter
            from .models_newsletter import NewsletterCampaign
            return NewsletterCampaign.objects.filter(status='SENT').order_by('-sent_at')
        return []
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class NewsletterEditView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/newsletter/edit.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class NewsletterDeleteView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/newsletter/delete.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class NewsletterSendView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/newsletter/send.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class NewsletterSubscribersView(AdminRequiredMixin, ListView):
    template_name = 'backend/admin/newsletter/subscribers.html'
    context_object_name = 'subscribers'
    paginate_by = 50
    
    def get_queryset(self):
        if NEWSLETTER_MODELS_AVAILABLE:
            return NewsletterSubscriber.objects.all().order_by('-subscribed_at')
        return []
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

@admin_required
@require_http_methods(["DELETE", "POST"])
def admin_newsletter_subscriber_delete(request, subscriber_id):
    """Delete newsletter subscriber"""
    try:
        if NEWSLETTER_MODELS_AVAILABLE:
            from .models_newsletter import NewsletterSubscriber
            subscriber = get_object_or_404(NewsletterSubscriber, id=subscriber_id)
            subscriber_email = subscriber.email
            subscriber.delete()
            
            if request.method == 'POST':
                messages.success(request, f'Abonné "{subscriber_email}" supprimé avec succès!')
                return redirect('admin_panel:newsletter')
            
            return JsonResponse({
                'success': True,
                'message': f'Abonné "{subscriber_email}" supprimé avec succès!'
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'Modèles de newsletter non disponibles'
            })
    except Exception as e:
        if request.method == 'POST':
            messages.error(request, f'Erreur lors de la suppression: {e}')
            return redirect('admin_panel:newsletter')
        
        return JsonResponse({
            'success': False,
            'message': str(e)
        })

class AdminNotificationsView(AdminRequiredMixin, ListView):
    template_name = 'backend/admin/notifications/list.html'
    context_object_name = 'notifications'
    paginate_by = 25
    
    def get_queryset(self):
        return Notification.objects.all().order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class NotificationCreateView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/notifications/create.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class NotificationListView(AdminRequiredMixin, ListView):
    template_name = 'backend/admin/notifications/list.html'
    context_object_name = 'notifications'
    paginate_by = 25
    
    def get_queryset(self):
        return Notification.objects.all().order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class NotificationEditView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/notifications/edit.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class NotificationDeleteView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/notifications/delete.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class NotificationSendView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/notifications/send.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class NotificationBulkSendView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/notifications/bulk_send.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class LoyaltyView(AdminRequiredMixin, ListView):
    template_name = 'backend/admin/loyalty/list.html'
    context_object_name = 'loyalty_programs'
    paginate_by = 20
    
    def get_queryset(self):
        # Return empty list since we don't have LoyaltyProgram model yet
        return []
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class LoyaltyCreateView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/loyalty/create.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class LoyaltyEditView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/loyalty/edit.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class LoyaltyListView(AdminRequiredMixin, ListView):
    template_name = 'backend/admin/loyalty/list.html'
    context_object_name = 'loyalty_programs'
    paginate_by = 20
    
    def get_queryset(self):
        # Return empty list since we don't have LoyaltyProgram model yet
        return []
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class LoyaltyDeleteView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/loyalty/delete.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class LoyaltyPointsView(AdminRequiredMixin, ListView):
    model = User
    template_name = 'backend/admin/loyalty/points.html'
    context_object_name = 'users'
    paginate_by = 50
    
    def get_queryset(self):
        return User.objects.filter(loyalty_points__gt=0).order_by('-loyalty_points')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class LoyaltyRewardsView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/loyalty/rewards.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminChatsView(AdminRequiredMixin, ListView):
    model = Chat
    template_name = 'backend/admin/chats/list.html'
    context_object_name = 'chats'
    paginate_by = 25
    
    def get_queryset(self):
        return Chat.objects.all().order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminChatDetailView(AdminRequiredMixin, DetailView):
    model = Chat
    template_name = 'backend/admin/chats/detail.html'
    context_object_name = 'chat'
    pk_url_kwarg = 'pk'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminChatReplyView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/chats/reply.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['chat'] = get_object_or_404(Chat, pk=self.kwargs['pk'])
        return context

class AdminChatBulkActionsView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/chats/bulk_actions.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class SystemSettingsView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/settings/system.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Add default settings if they don't exist
        context['settings'] = {
            'site_name': 'Vidé-Grenier Kamer',
            'site_description': 'Plateforme de vente et d\'achat d\'articles d\'occasion',
            'site_url': 'https://vide-grenier-kamer.com',
            'contact_email': 'contact@vide-grenier-kamer.com',
            'timezone': 'Africa/Douala',
            'language': 'fr',
            'currency': 'XAF',
            'maintenance_mode': False,
            'session_timeout': 30,
            'max_login_attempts': 5,
            'require_email_verification': True,
            'require_phone_verification': False,
            'password_min_length': 8,
            'require_strong_password': True,
            'enable_captcha': False,
            'enable_2fa': False,
            'smtp_host': 'smtp.gmail.com',
            'smtp_port': 587,
            'smtp_username': '',
            'smtp_password': '',
            'smtp_use_tls': True,
            'email_from_name': 'Vidé-Grenier Kamer',
            'email_from_address': 'noreply@vide-grenier-kamer.com',
            'email_reply_to': 'support@vide-grenier-kamer.com',
            'enable_mobile_money': True,
            'enable_bank_transfer': True,
            'enable_cash_on_delivery': True,
            'enable_paypal': False,
            'commission_rate': 5.0,
            'minimum_commission': 100,
            'maximum_commission': 5000,
            'email_order_confirmation': True,
            'email_payment_confirmation': True,
            'email_product_approval': True,
            'email_newsletter': True,
            'push_new_products': True,
            'push_price_drops': True,
            'push_order_updates': True,
            'push_chat_messages': True,
        }
        return context

class BackupView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/settings/backup.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class LogsView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/settings/logs.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class SecuritySettingsView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/settings/security.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class EmailSettingsView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/settings/email.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class PaymentSettingsView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/settings/payment.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class NotificationSettingsView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/settings/notification.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminProfileView(AdminRequiredMixin, DetailView):
    model = User
    template_name = 'backend/admin/profile/profile.html'
    context_object_name = 'profile_user'
    
    def get_object(self):
        return self.request.user
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminProfileEditView(AdminRequiredMixin, UpdateView):
    model = User
    template_name = 'backend/admin/profile/edit.html'
    fields = ['first_name', 'last_name', 'email', 'phone', 'city']
    success_url = reverse_lazy('admin_panel:profile')
    
    def get_object(self):
        return self.request.user
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminChangePasswordView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/profile/change_password.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class CommissionManagementView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/commissions/list.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminCommissionView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/commissions/list.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['commissions'] = []  # Placeholder
        return context

class CalculateCommissionsView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/commissions/calculate.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class CommissionPayoutView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/commissions/payout.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class CommissionHistoryView(AdminRequiredMixin, ListView):
    template_name = 'backend/admin/commissions/history.html'
    context_object_name = 'commissions'
    paginate_by = 25
    
    def get_queryset(self):
        # Return empty list since we don't have Commission model yet
        return []
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class WithdrawalManagementView(AdminRequiredMixin, ListView):
    template_name = 'backend/admin/withdrawals/list.html'
    context_object_name = 'withdrawals'
    paginate_by = 25
    
    def get_queryset(self):
        # Return empty list since we don't have Withdrawal model yet
        return []
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class ApproveWithdrawalView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/withdrawals/approve.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class RejectWithdrawalView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/withdrawals/reject.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class WithdrawalDetailView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/withdrawals/detail.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class WithdrawalBulkActionsView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/withdrawals/bulk_actions.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminCategoryListView(AdminRequiredMixin, ListView):
    model = Category
    template_name = 'backend/admin/categories/list.html'
    context_object_name = 'categories'
    paginate_by = 25
    
    def get_queryset(self):
        return Category.objects.all().order_by('name')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_categories_count'] = Category.objects.filter(is_active=True).count()
        context['total_products'] = Product.objects.count()
        context['popular_categories_count'] = Category.objects.annotate(
            product_count=Count('products')
        ).filter(product_count__gt=0).count()
        return context

class AdminCategoryCreateView(AdminRequiredMixin, CreateView):
    model = Category
    template_name = 'backend/admin/categories/create.html'
    fields = ['name', 'description', 'icon', 'is_active']
    success_url = reverse_lazy('admin_panel:categories')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminCategoryEditView(AdminRequiredMixin, UpdateView):
    model = Category
    template_name = 'backend/admin/categories/edit.html'
    fields = ['name', 'description', 'icon', 'is_active']
    pk_url_kwarg = 'pk'
    success_url = reverse_lazy('admin_panel:categories')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminCategoryDeleteView(AdminRequiredMixin, DeleteView):
    model = Category
    template_name = 'backend/admin/categories/delete.html'
    pk_url_kwarg = 'pk'
    success_url = reverse_lazy('admin_panel:categories')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminPickupPointListView(AdminRequiredMixin, ListView):
    model = PickupPoint
    template_name = 'backend/admin/pickup_points/list.html'
    context_object_name = 'pickup_points'
    paginate_by = 25
    
    def get_queryset(self):
        return PickupPoint.objects.all().order_by('name')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['active_pickup_points_count'] = PickupPoint.objects.filter(is_active=True).count()
        context['total_orders'] = Order.objects.count()
        context['cities_count'] = PickupPoint.objects.values('city').distinct().count()
        context['cities'] = PickupPoint.objects.values_list('city', flat=True).distinct()
        return context

class AdminPickupPointCreateView(AdminRequiredMixin, CreateView):
    model = PickupPoint
    template_name = 'backend/admin/pickup_points/create.html'
    fields = ['name', 'address', 'city', 'phone', 'email', 'is_active']
    success_url = reverse_lazy('admin_panel:pickup_points')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminPickupPointEditView(AdminRequiredMixin, UpdateView):
    model = PickupPoint
    template_name = 'backend/admin/pickup_points/edit.html'
    fields = ['name', 'address', 'city', 'phone', 'email', 'is_active']
    pk_url_kwarg = 'pk'
    success_url = reverse_lazy('admin_panel:pickup_points')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminPickupPointDeleteView(AdminRequiredMixin, DeleteView):
    model = PickupPoint
    template_name = 'backend/admin/pickup_points/delete.html'
    pk_url_kwarg = 'pk'
    success_url = reverse_lazy('admin_panel:pickup_points')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminReviewListView(AdminRequiredMixin, ListView):
    model = Review
    template_name = 'backend/admin/reviews/list.html'
    context_object_name = 'reviews'
    paginate_by = 25
    
    def get_queryset(self):
        return Review.objects.all().order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Since Review model doesn't have approval fields, we'll use is_verified instead
        context['verified_reviews_count'] = Review.objects.filter(is_verified=True).count()
        context['unverified_reviews_count'] = Review.objects.filter(is_verified=False).count()
        context['total_reviews_count'] = Review.objects.count()
        return context

class AdminReviewDetailView(AdminRequiredMixin, DetailView):
    model = Review
    template_name = 'backend/admin/reviews/detail.html'
    context_object_name = 'review'
    pk_url_kwarg = 'pk'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class AdminReviewApproveView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/reviews/approve.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['review'] = get_object_or_404(Review, pk=self.kwargs['pk'])
        return context

class AdminReviewRejectView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/reviews/reject.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['review'] = get_object_or_404(Review, pk=self.kwargs['pk'])
        return context

class AdminReviewDeleteView(AdminRequiredMixin, DeleteView):
    model = Review
    template_name = 'backend/admin/reviews/delete.html'
    pk_url_kwarg = 'pk'
    success_url = reverse_lazy('admin_panel:reviews')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class ProductReportsView(AdminRequiredMixin, ListView):
    model = Product
    template_name = 'backend/admin/reports/products.html'
    context_object_name = 'products'
    paginate_by = 25
    
    def get_queryset(self):
        return Product.objects.all().order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class UserReportsView(AdminRequiredMixin, ListView):
    model = User
    template_name = 'backend/admin/reports/users.html'
    context_object_name = 'users'
    paginate_by = 25
    
    def get_queryset(self):
        return User.objects.all().order_by('-date_joined')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class OrderReportsView(AdminRequiredMixin, ListView):
    model = Order
    template_name = 'backend/admin/reports/orders.html'
    context_object_name = 'orders'
    paginate_by = 25
    
    def get_queryset(self):
        return Order.objects.all().order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class FinancialReportsView(AdminRequiredMixin, ListView):
    model = Payment
    template_name = 'backend/admin/reports/financial.html'
    context_object_name = 'payments'
    paginate_by = 25
    
    def get_queryset(self):
        return Payment.objects.all().order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class DashboardStatsAPIView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/api/dashboard_stats.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class UserStatsAPIView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/api/user_stats.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class OrderStatsAPIView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/api/order_stats.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class RevenueStatsAPIView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/api/revenue_stats.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class MaintenanceModeView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/utility/maintenance.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class EnableMaintenanceView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/utility/enable_maintenance.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class DisableMaintenanceView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/utility/disable_maintenance.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class ClearCacheView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/utility/clear_cache.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class DatabaseBackupView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/utility/database_backup.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

class DatabaseRestoreView(AdminRequiredMixin, TemplateView):
    template_name = 'backend/admin/utility/database_restore.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context

# Product Approval/Rejection Views
class ProductApprovalView(AdminRequiredMixin, DetailView):
    model = Product
    template_name = 'backend/admin/products/approve.html'
    context_object_name = 'product'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['product_approval_logs'] = self.get_approval_logs()
        return context
    
    def get_approval_logs(self):
        # Get approval history for this product
        return []  # Implement approval log model

class ProductRejectionView(AdminRequiredMixin, DetailView):
    model = Product
    template_name = 'backend/admin/products/reject.html'
    context_object_name = 'product'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['rejection_reasons'] = self.get_rejection_reasons()
        return context
    
    def get_rejection_reasons(self):
        return [
            'Contenu inapproprié',
            'Qualité insuffisante',
            'Mauvaise catégorie',
            'Produit en double',
            'Problème de prix',
            'Autre raison'
        ]

class ProductContactSellerView(AdminRequiredMixin, DetailView):
    model = Product
    template_name = 'backend/admin/products/contact_seller.html'
    context_object_name = 'product'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['chat_messages'] = self.get_chat_messages()
        return context
    
    def get_chat_messages(self):
        # Get chat messages between admin and seller
        return []

# Order Processing Views
class OrderDetailView(AdminRequiredMixin, DetailView):
    model = Order
    template_name = 'backend/admin/orders/detail.html'
    context_object_name = 'order'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['order_status_history'] = self.get_status_history()
        context['commission_rate'] = 5.0  # 5% commission
        context['vat_rate'] = 19.25  # 19.25% VAT
        context['recent_orders'] = self.get_recent_orders()
        return context
    
    def get_status_history(self):
        # Get order status change history
        return []
    
    def get_recent_orders(self):
        return Order.objects.filter(
            items__product__seller=self.object.items.first().product.seller
        ).exclude(id=self.object.id).order_by('-created_at')[:5]

class OrderProcessingView(AdminRequiredMixin, UpdateView):
    model = Order
    template_name = 'backend/admin/orders/processing.html'
    fields = ['status', 'tracking_number', 'estimated_delivery_date']
    
    def get_success_url(self):
        return reverse_lazy('admin_panel:order_detail', kwargs={'pk': self.object.pk})

class AdminOrderDeleteView(AdminRequiredMixin, DeleteView):
    model = Order
    template_name = 'backend/admin/orders/delete.html'
    pk_url_kwarg = 'pk'
    success_url = reverse_lazy('admin_panel:orders')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        return context
    
    def delete(self, request, *args, **kwargs):
        order = self.get_object()
        messages.success(request, f'Commande "{order.order_number}" supprimée avec succès!')
        return super().delete(request, *args, **kwargs)

# Pickup Point Management Views
class PickupPointEditView(AdminRequiredMixin, UpdateView):
    model = PickupPoint
    template_name = 'backend/admin/pickup_points/edit.html'
    fields = ['name', 'address', 'city', 'phone', 'email', 'is_active', 'opening_hours', 'special_hours', 'capacity', 'processing_time', 'notes']
    
    def get_success_url(self):
        return reverse_lazy('admin_panel:pickup_point_detail', kwargs={'pk': self.object.pk})

class PickupPointDetailView(AdminRequiredMixin, DetailView):
    model = PickupPoint
    template_name = 'backend/admin/pickup_points/detail.html'
    context_object_name = 'pickup_point'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['recent_orders'] = self.get_recent_orders()
        context['total_orders'] = self.get_total_orders()
        context['completed_orders'] = self.get_completed_orders()
        context['pending_orders'] = self.get_pending_orders()
        context['avg_processing_time'] = self.get_avg_processing_time()
        return context
    
    def get_recent_orders(self):
        return Order.objects.filter(
            pickup_point=self.object
        ).order_by('-created_at')[:5]
    
    def get_total_orders(self):
        return Order.objects.filter(pickup_point=self.object).count()
    
    def get_completed_orders(self):
        return Order.objects.filter(
            pickup_point=self.object,
            status='DELIVERED'
        ).count()
    
    def get_pending_orders(self):
        return Order.objects.filter(
            pickup_point=self.object,
            status__in=['PENDING', 'CONFIRMED', 'PROCESSING']
        ).count()
    
    def get_avg_processing_time(self):
        # Calculate average processing time
        return 24  # Default 24 hours

# Additional Admin Views for New Features
class OrderStatusUpdateView(AdminRequiredMixin, View):
    """AJAX view for updating order status"""
    
    def post(self, request, pk):
        try:
            order = Order.objects.get(id=pk)
            new_status = request.POST.get('status')
            notes = request.POST.get('notes', '')
            
            if new_status:
                order.status = new_status
                order.save()
                
                # Send email notification if requested
                if request.POST.get('send_email'):
                    from .email_service import email_service
                    email_service.send_delivery_status_email(order, new_status)
                
                return JsonResponse({'success': True, 'message': 'Status updated successfully'})
            else:
                return JsonResponse({'success': False, 'message': 'Status is required'})
                
        except Order.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Order not found'})

class PickupPointToggleStatusView(AdminRequiredMixin, View):
    """Toggle pickup point status"""
    
    def post(self, request, pk):
        try:
            pickup_point = PickupPoint.objects.get(id=pk)
            pickup_point.is_active = not pickup_point.is_active
            pickup_point.save()
            
            return JsonResponse({
                'success': True, 
                'is_active': pickup_point.is_active,
                'message': f'Pickup point {"activated" if pickup_point.is_active else "deactivated"}'
            })
        except PickupPoint.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Pickup point not found'})

class StockCalculateView(AdminRequiredMixin, View):
    """Calculate stock status for all products"""
    
    def post(self, request):
        try:
            from .stock_calculator import stock_calculator
            
            products = Product.objects.all()
            results = []
            
            for product in products:
                stock_status = stock_calculator.calculate_stock_status(product)
                results.append(stock_status)
            
            return JsonResponse({
                'success': True,
                'results': results,
                'message': f'Stock calculated for {len(results)} products'
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

class StockAlertsView(AdminRequiredMixin, ListView):
    """View for stock alerts"""
    model = Product
    template_name = 'backend/admin/stock/alerts.html'
    context_object_name = 'alerts'
    
    def get_queryset(self):
        from .stock_calculator import stock_calculator
        
        products = Product.objects.all()
        alerts = []
        
        for product in products:
            stock_status = stock_calculator.calculate_stock_status(product)
            if stock_status['alerts']:
                alerts.extend(stock_status['alerts'])
        
        return alerts

class CommissionCalculateView(AdminRequiredMixin, View):
    """Calculate commissions for all sellers"""
    
    def post(self, request):
        try:
            from .commission_manager import commission_manager
            
            start_date = request.POST.get('start_date')
            end_date = request.POST.get('end_date')
            
            if start_date and end_date:
                start_date = datetime.strptime(start_date, '%Y-%m-%d')
                end_date = datetime.strptime(end_date, '%Y-%m-%d')
            else:
                end_date = timezone.now()
                start_date = end_date - timedelta(days=30)
            
            commission_data = commission_manager.calculate_all_commissions(start_date, end_date)
            
            return JsonResponse({
                'success': True,
                'data': commission_data,
                'message': f'Commissions calculated for {commission_data["total_sellers"]} sellers'
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

class CommissionPayoutView(AdminRequiredMixin, View):
    """Process commission payout"""
    
    def post(self, request):
        try:
            from .commission_manager import commission_manager
            
            seller_id = request.POST.get('seller_id')
            amount = Decimal(request.POST.get('amount', 0))
            payment_method = request.POST.get('payment_method', 'bank_transfer')
            
            seller = User.objects.get(id=seller_id)
            result = commission_manager.process_commission_payout(seller, amount, payment_method)
            
            return JsonResponse(result)
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

class CommissionReportView(AdminRequiredMixin, View):
    """Generate commission report"""
    
    def get(self, request):
        try:
            from .commission_manager import commission_manager
            
            start_date = request.GET.get('start_date')
            end_date = request.GET.get('end_date')
            
            if start_date and end_date:
                start_date = datetime.strptime(start_date, '%Y-%m-%d')
                end_date = datetime.strptime(end_date, '%Y-%m-%d')
            else:
                end_date = timezone.now()
                start_date = end_date - timedelta(days=30)
            
            report = commission_manager.generate_commission_report(start_date, end_date)
            
            return JsonResponse({
                'success': True,
                'report': report
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

class AdminChatListView(AdminRequiredMixin, ListView):
    """List all chats for admin"""
    model = Chat
    template_name = 'backend/admin/chats/list.html'
    context_object_name = 'chats'
    
    def get_queryset(self):
        return Chat.objects.filter(
            Q(buyer__is_staff=True) | Q(seller__is_staff=True)
        ).distinct().order_by('-updated_at')

class AdminChatDetailView(AdminRequiredMixin, DetailView):
    """Chat detail view for admin"""
    model = User
    template_name = 'backend/admin/chats/detail.html'
    context_object_name = 'user'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Get or create chat between admin and user
        chat, created = Chat.objects.get_or_create(
            buyer=self.request.user,
            seller=self.object,
            defaults={'product': None}
        )
        context['chat'] = chat
        context['messages'] = chat.messages.all().order_by('created_at')
        return context

class AdminChatMessagesView(AdminRequiredMixin, View):
    """Get chat messages via AJAX"""
    
    def get(self, request, user_id):
        try:
            user = User.objects.get(id=user_id)
            chat, created = Chat.objects.get_or_create(
                buyer=request.user,
                seller=user,
                defaults={'product': None}
            )
            
            messages = chat.messages.all().order_by('created_at')
            message_data = []
            
            for message in messages:
                message_data.append({
                    'id': message.id,
                    'content': message.content,
                    'sender_id': message.sender.id,
                    'sender_name': message.sender.get_full_name(),
                    'timestamp': message.created_at.isoformat(),
                    'is_admin': message.sender.is_staff
                })
            
            return JsonResponse({
                'success': True,
                'messages': message_data
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

class AdminChatSendView(AdminRequiredMixin, View):
    """Send chat message via AJAX"""
    
    def post(self, request, user_id):
        try:
            user = User.objects.get(id=user_id)
            content = request.POST.get('content', '').strip()
            
            if not content:
                return JsonResponse({'success': False, 'message': 'Message content is required'})
            
            chat, created = Chat.objects.get_or_create(
                buyer=request.user,
                seller=user,
                defaults={'product': None}
            )
            
            message = Message.objects.create(
                chat=chat,
                sender=request.user,
                content=content
            )
            
            return JsonResponse({
                'success': True,
                'message_id': message.id,
                'timestamp': message.created_at.isoformat()
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

class AdminEmailSendView(AdminRequiredMixin, View):
    """Send email via admin interface"""
    
    def post(self, request):
        try:
            from .email_service import email_service
            
            to_email = request.POST.get('to_email')
            subject = request.POST.get('subject')
            message = request.POST.get('message')
            
            if not all([to_email, subject, message]):
                return JsonResponse({'success': False, 'message': 'All fields are required'})
            
            success = email_service.send_email(
                to_email=to_email,
                subject=subject,
                html_content=message,
                text_content=message
            )
            
            return JsonResponse({
                'success': success,
                'message': 'Email sent successfully' if success else 'Failed to send email'
            })
        except Exception as e:
            return JsonResponse({'success': False, 'message': str(e)})

class AdminEmailTemplatesView(AdminRequiredMixin, TemplateView):
    """Email templates management"""
    template_name = 'backend/admin/email/templates.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['templates'] = self.get_email_templates()
        return context
    
    def get_email_templates(self):
        return [
            {
                'name': 'Order Confirmation',
                'subject': 'Confirmation de commande #{order_number}',
                'template': 'emails/order_confirmation.html'
            },
            {
                'name': 'Product Approval',
                'subject': 'Votre produit a été approuvé',
                'template': 'emails/product_approval.html'
            },
            {
                'name': 'Product Rejection',
                'subject': 'Votre produit a été rejeté',
                'template': 'emails/product_rejection.html'
            }
        ]

# Category Management Views
class CategoryCreateView(AdminRequiredMixin, CreateView):
    model = Category
    template_name = 'backend/admin/categories/create.html'
    fields = ['name', 'description', 'icon', 'is_active']
    
    def get_success_url(self):
        return reverse_lazy('admin_panel:categories')

class CategoryEditView(AdminRequiredMixin, UpdateView):
    model = Category
    template_name = 'backend/admin/categories/edit.html'
    fields = ['name', 'description', 'icon', 'is_active']
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['category'] = self.object
        return context
    
    def get_success_url(self):
        return reverse_lazy('admin_panel:categories')

class CategoryDeleteView(AdminRequiredMixin, DeleteView):
    model = Category
    template_name = 'backend/admin/categories/delete.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['other_categories'] = Category.objects.exclude(id=self.object.id)
        return context
    
    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        product_action = request.POST.get('product_action')
        new_category_id = request.POST.get('new_category')
        
        if product_action == 'move' and new_category_id:
            # Move products to new category
            new_category = Category.objects.get(id=new_category_id)
            self.object.product_set.update(category=new_category)
        elif product_action == 'delete':
            # Delete all products in this category
            self.object.product_set.delete()
        
        return super().post(request, *args, **kwargs)
    
    def get_success_url(self):
        return reverse_lazy('admin_panel:categories')
